import requests
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import pytz 
from io import BytesIO
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import time
import json
import os
import pypdf
import pandas as pd
import re
import shutil
import numpy as np
from pymongo import MongoClient 
from collections import defaultdict

from flask import Flask, request, render_template_string, send_file, flash, session, redirect, url_for, make_response, jsonify

app = Flask(__name__)
app.secret_key = 'super-secret-secure-key-bd' 

# ==============================================================================
# কনফিগারেশন
# ==============================================================================

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=5) 

bd_tz = pytz.timezone('Asia/Dhaka')

def get_bd_time():
    return datetime.now(bd_tz)

def get_bd_date_str():
    return get_bd_time().strftime('%d-%m-%Y')

# ==============================================================================
# Browser Cache Control
# ==============================================================================
@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# ==============================================================================
# MongoDB Connection
# ==============================================================================
MONGO_URI = "mongodb+srv://Mehedi:Mehedi123@office.jxdnuaj.mongodb.net/?appName=Office"

try:
    client = MongoClient(MONGO_URI)
    db = client['office_db']
    users_col = db['users']
    stats_col = db['stats']
    accessories_col = db['accessories']
    print("MongoDB Connected Successfully!")
except Exception as e:
    print(f"MongoDB Connection Error: {e}")
    # ==============================================================================
# COMMON CSS STYLES
# ==============================================================================

COMMON_STYLES = """
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/animate.css/4.1.1/animate.min.css">
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        :root {
            --bg-body: #0a0a0f;
            --bg-sidebar: #12121a;
            --bg-card: #16161f;
            --bg-card-hover: #1a1a25;
            --text-primary: #FFFFFF;
            --text-secondary: #8b8b9e;
            --accent-orange: #FF7A00;
            --accent-orange-light: #FF9A40;
            --accent-orange-dark: #E56D00;
            --accent-orange-glow: rgba(255, 122, 0, 0.3);
            --accent-purple: #8B5CF6;
            --accent-green: #10B981;
            --accent-red: #EF4444;
            --accent-blue: #3B82F6;
            --accent-cyan: #06B6D4;
            --border-color: rgba(255, 255, 255, 0.08);
            --border-glow: rgba(255, 122, 0, 0.2);
            --card-radius: 16px;
            --transition-smooth: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            --shadow-card: 0 4px 24px rgba(0, 0, 0, 0.4);
            --shadow-glow: 0 0 40px rgba(255, 122, 0, 0.15);
            --gradient-orange: linear-gradient(135deg, #FF7A00 0%, #FF9A40 100%);
            --gradient-dark: linear-gradient(180deg, #12121a 0%, #0a0a0f 100%);
            --gradient-card: linear-gradient(145deg, rgba(22, 22, 31, 0.9) 0%, rgba(16, 16, 22, 0.95) 100%);
        }

        * { 
            margin: 0;
            padding: 0; 
            box-sizing: border-box; 
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; 
        }
        
        ::-webkit-scrollbar { width: 6px; height: 6px; }
        ::-webkit-scrollbar-track { background: var(--bg-body); }
        ::-webkit-scrollbar-thumb { background: var(--accent-orange); border-radius: 10px; }
        ::-webkit-scrollbar-thumb:hover { background: var(--accent-orange-light); }
        
        body {
            background: var(--bg-body);
            color: var(--text-primary);
            min-height: 100vh;
            display: flex;
            overflow-x: hidden;
            position: relative;
        }

        .animated-bg {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: 
                radial-gradient(ellipse at 20% 20%, rgba(255, 122, 0, 0.08) 0%, transparent 50%),
                radial-gradient(ellipse at 80% 80%, rgba(139, 92, 246, 0.06) 0%, transparent 50%),
                radial-gradient(ellipse at 50% 50%, rgba(16, 185, 129, 0.04) 0%, transparent 70%);
            z-index: 0;
            animation: bgPulse 15s ease-in-out infinite;
        }

        @keyframes bgPulse {
            0%, 100% { opacity: 1; transform: scale(1); }
            50% { opacity: 0.8; transform: scale(1.05); }
        }

        /* Sidebar */
        .sidebar {
            width: 280px;
            height: 100vh; 
            background: var(--gradient-dark);
            position: fixed; 
            top: 0; 
            left: 0; 
            display: flex; 
            flex-direction: column;
            padding: 30px 20px;
            border-right: 1px solid var(--border-color); 
            z-index: 1000;
            transition: var(--transition-smooth);
            box-shadow: 4px 0 30px rgba(0, 0, 0, 0.3);
        }
        
        .sidebar::before {
            content: '';
            position: absolute;
            top: 0;
            right: 0;
            width: 1px;
            height: 100%;
            background: linear-gradient(180deg, transparent, var(--accent-orange), transparent);
            opacity: 0.3;
        }

        .brand-logo { 
            font-size: 26px;
            font-weight: 900; 
            color: white; 
            margin-bottom: 50px; 
            display: flex; 
            align-items: center; 
            gap: 12px; 
            padding: 0 10px;
        }

        .brand-logo span { 
            background: var(--gradient-orange);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .brand-logo i {
            font-size: 28px;
            color: var(--accent-orange);
            filter: drop-shadow(0 0 10px var(--accent-orange-glow));
            animation: logoFloat 3s ease-in-out infinite;
        }

        @keyframes logoFloat {
            0%, 100% { transform: translateY(0) rotate(0deg); }
            50% { transform: translateY(-5px) rotate(5deg); }
        }
        
        .nav-menu { 
            flex-grow: 1; 
            display: flex; 
            flex-direction: column;
            gap: 8px; 
        }

        .nav-link {
            display: flex;
            align-items: center; 
            padding: 14px 18px; 
            color: var(--text-secondary);
            text-decoration: none; 
            border-radius: 12px; 
            transition: var(--transition-smooth);
            cursor: pointer; 
            font-weight: 500; 
            font-size: 14px;
            position: relative;
            overflow: hidden;
        }

        .nav-link::before {
            content: '';
            position: absolute;
            left: 0;
            top: 0;
            width: 0;
            height: 100%;
            background: linear-gradient(90deg, var(--accent-orange-glow), transparent);
            transition: var(--transition-smooth);
            z-index: -1;
        }
        
        .nav-link:hover::before, .nav-link.active::before { width: 100%; }

        .nav-link:hover, .nav-link.active { 
            color: var(--accent-orange); 
            transform: translateX(5px);
        }

        .nav-link.active {
            background: rgba(255, 122, 0, 0.1);
            border-left: 3px solid var(--accent-orange);
            box-shadow: 0 0 20px var(--accent-orange-glow);
        }

        .nav-link i { 
            width: 24px;
            margin-right: 12px; 
            font-size: 18px; 
            text-align: center;
            transition: var(--transition-smooth);
        }

        .nav-link:hover i {
            transform: scale(1.2);
            filter: drop-shadow(0 0 8px var(--accent-orange));
        }

        .nav-link .nav-badge {
            margin-left: auto;
            background: var(--accent-orange);
            color: white;
            padding: 2px 8px;
            border-radius: 10px;
            font-size: 11px;
            font-weight: 700;
        }

        .sidebar-footer {
            margin-top: auto;
            padding-top: 20px; 
            border-top: 1px solid var(--border-color);
            text-align: center; 
            font-size: 11px; 
            color: var(--text-secondary); 
            font-weight: 500; 
            opacity: 0.5;
            letter-spacing: 1px;
        }

        /* Main Content */
        .main-content { 
            margin-left: 280px;
            width: calc(100% - 280px); 
            padding: 30px 40px; 
            position: relative;
            z-index: 1;
            min-height: 100vh;
        }

        .header-section { 
            display: flex;
            justify-content: space-between; 
            align-items: flex-start; 
            margin-bottom: 35px;
            animation: fadeInDown 0.6s ease-out;
        }

        @keyframes fadeInDown {
            from { opacity: 0; transform: translateY(-20px); }
            to { opacity: 1; transform: translateY(0); }
        }

        .page-title { 
            font-size: 32px;
            font-weight: 800; 
            color: white; 
            margin-bottom: 8px;
            letter-spacing: -0.5px;
            background: linear-gradient(135deg, #fff 0%, #ccc 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .page-subtitle { 
            color: var(--text-secondary);
            font-size: 14px;
            font-weight: 400;
        }

        .status-badge {
            background: var(--bg-card);
            padding: 12px 24px;
            border-radius: 50px;
            border: 1px solid var(--border-color);
            font-size: 13px;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 10px;
            box-shadow: var(--shadow-card);
        }
        
        .status-dot {
            width: 10px;
            height: 10px;
            background: var(--accent-green);
            border-radius: 50%;
            animation: statusPulse 1.5s ease-in-out infinite;
            box-shadow: 0 0 10px var(--accent-green);
        }
        
        @keyframes statusPulse {
            0%, 100% { opacity: 1; transform: scale(1); }
            50% { opacity: 0.8; transform: scale(1.2); }
        }

        /* Cards & Grid */
        .stats-grid { 
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); 
            gap: 24px; 
            margin-bottom: 35px;
        }

        .dashboard-grid-2 { 
            display: grid;
            grid-template-columns: 2fr 1fr; 
            gap: 24px; 
            margin-bottom: 24px; 
        }

        .card { 
            background: var(--gradient-card);
            border: 1px solid var(--border-color);
            border-radius: var(--card-radius);
            padding: 28px;
            backdrop-filter: blur(10px);
            transition: var(--transition-smooth);
            position: relative;
            overflow: hidden;
        }

        .card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 1px;
            background: linear-gradient(90deg, transparent, var(--accent-orange-glow), transparent);
            opacity: 0;
            transition: var(--transition-smooth);
        }

        .card:hover {
            border-color: var(--border-glow);
            box-shadow: var(--shadow-glow);
            transform: translateY(-4px);
        }

        .card:hover::before { opacity: 1; }
        
        .section-header { 
            display: flex;
            justify-content: space-between; 
            align-items: center; 
            margin-bottom: 24px; 
            font-weight: 700;
            font-size: 16px; 
            color: white;
            letter-spacing: -0.3px;
        }

        /* Stat Cards */
        .stat-card { 
            display: flex;
            align-items: center; 
            gap: 24px; 
            transition: var(--transition-smooth);
            cursor: pointer;
        }

        .stat-card:hover { 
            transform: translateY(-6px) scale(1.02);
        }

        .stat-icon { 
            width: 64px;
            height: 64px; 
            background: linear-gradient(145deg, rgba(255, 122, 0, 0.15), rgba(255, 122, 0, 0.05));
            border-radius: 16px; 
            display: flex; 
            justify-content: center; 
            align-items: center;
            font-size: 26px; 
            color: var(--accent-orange);
            transition: var(--transition-smooth);
        }

        .stat-card:hover .stat-icon {
            transform: rotate(-10deg) scale(1.1);
            box-shadow: 0 0 30px var(--accent-orange-glow);
        }

        .stat-info h3 { 
            font-size: 36px;
            font-weight: 800; 
            margin: 0; 
            color: white;
            letter-spacing: -1px;
            line-height: 1;
        }

        .stat-info p { 
            font-size: 13px;
            color: var(--text-secondary); 
            margin: 6px 0 0 0; 
            text-transform: uppercase;
            letter-spacing: 1.5px;
            font-weight: 600;
        }

        /* Progress Bars */
        .progress-item { margin-bottom: 24px; }

        .progress-header {
            display: flex;
            justify-content: space-between;
            margin-bottom: 10px;
            font-size: 14px;
            color: white;
            font-weight: 500;
        }

        .progress-value {
            color: var(--text-secondary);
            font-weight: 600;
        }

        .progress-bar-container {
            height: 8px;
            background: rgba(255, 255, 255, 0.05);
            border-radius: 10px;
            overflow: hidden;
        }

        .progress-bar-fill {
            height: 100%;
            border-radius: 10px;
            animation: progressFill 1.5s ease-out forwards;
            transform-origin: left;
        }

        @keyframes progressFill {
            from { transform: scaleX(0); }
            to { transform: scaleX(1); }
        }

        .progress-orange { background: var(--gradient-orange); }
        .progress-purple { background: linear-gradient(135deg, #8B5CF6 0%, #A78BFA 100%); }
        .progress-green { background: linear-gradient(135deg, #10B981 0%, #34D399 100%); }

        /* Forms */
        .input-group { margin-bottom: 20px; }

        .input-group label { 
            display: block;
            font-size: 11px; 
            color: var(--text-secondary); 
            margin-bottom: 8px; 
            text-transform: uppercase; 
            font-weight: 700;
            letter-spacing: 1.5px;
        }

        input, select { 
            width: 100%;
            padding: 14px 18px; 
            background: rgba(255, 255, 255, 0.03);
            border: 1px solid var(--border-color);
            border-radius: 12px; 
            color: white; 
            font-size: 15px; 
            font-weight: 500;
            outline: none; 
            transition: var(--transition-smooth);
        }

        input::placeholder {
            color: var(--text-secondary);
            opacity: 0.5;
        }

        input:focus, select:focus { 
            border-color: var(--accent-orange);
            background: rgba(255, 122, 0, 0.05);
            box-shadow: 0 0 0 4px var(--accent-orange-glow);
        }

        select {
            cursor: pointer;
            appearance: none;
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' fill='%23FF7A00' viewBox='0 0 24 24'%3E%3Cpath d='M7 10l5 5 5-5z'/%3E%3C/svg%3E");
            background-repeat: no-repeat;
            background-position: right 12px center;
            background-size: 24px;
            background-color: rgba(255, 255, 255, 0.03);
        }
        
        select option {
            background-color: #1a1a25;
            color: white;
            padding: 10px;
        }
        
        button { 
            width: 100%;
            padding: 14px 24px; 
            background: var(--gradient-orange);
            color: white; 
            border: none; 
            border-radius: 12px; 
            font-weight: 700;
            font-size: 15px;
            cursor: pointer; 
            transition: var(--transition-smooth);
            position: relative;
            overflow: hidden;
            letter-spacing: 0.5px;
        }
        
        button::before {
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent);
            transition: 0.5s;
        }

        button:hover::before { left: 100%; }

        button:hover { 
            transform: translateY(-3px);
            box-shadow: 0 10px 30px var(--accent-orange-glow);
        }

        /* Tables */
        .dark-table { 
            width: 100%;
            border-collapse: collapse; 
            margin-top: 10px; 
        }

        .dark-table th { 
            text-align: left;
            padding: 14px 16px; 
            color: var(--text-secondary); 
            font-size: 11px;
            text-transform: uppercase;
            letter-spacing: 1.5px;
            border-bottom: 1px solid var(--border-color);
            font-weight: 700;
        }

        .dark-table td { 
            padding: 16px;
            color: white; 
            font-size: 14px; 
            border-bottom: 1px solid rgba(255,255,255,0.03);
            vertical-align: middle;
            transition: var(--transition-smooth);
        }

        .dark-table tr:hover td { 
            background: rgba(255, 122, 0, 0.03);
        }

        .table-badge {
            background: rgba(255,255,255,0.05);
            padding: 6px 14px;
            border-radius: 8px;
            font-size: 12px;
            font-weight: 600;
            display: inline-block;
        }
        
        /* Action Buttons */
        .action-cell { 
            display: flex;
            gap: 8px; 
            justify-content: flex-end; 
        }

        .action-btn { 
            padding: 8px 12px;
            border-radius: 8px; 
            text-decoration: none; 
            font-size: 12px; 
            display: inline-flex; 
            align-items: center; 
            justify-content: center; 
            cursor: pointer; 
            border: none; 
            transition: var(--transition-smooth);
        }

        .btn-edit { 
            background: rgba(139, 92, 246, 0.15);
            color: #A78BFA; 
        }

        .btn-edit:hover { 
            background: var(--accent-purple);
            color: white;
            transform: scale(1.1);
        }

        .btn-del { 
            background: rgba(239, 68, 68, 0.15);
            color: #F87171; 
        }

        .btn-del:hover { 
            background: var(--accent-red);
            color: white;
            transform: scale(1.1);
        }

        /* Loading Overlay */
        #loading-overlay { 
            display: none;
            position: fixed; 
            top: 0; 
            left: 0; 
            width: 100%; 
            height: 100%; 
            background: rgba(10, 10, 15, 0.95);
            z-index: 9999; 
            flex-direction: column;
            justify-content: center; 
            align-items: center; 
            backdrop-filter: blur(20px);
        }
        
        .spinner-container {
            position: relative;
            width: 80px;
            height: 80px;
        }

        .spinner { 
            width: 80px;
            height: 80px; 
            border: 4px solid rgba(255, 122, 0, 0.1);
            border-top: 4px solid var(--accent-orange);
            border-right: 4px solid var(--accent-orange-light);
            border-radius: 50%;
            animation: spin 0.8s linear infinite;
            box-shadow: 0 0 30px var(--accent-orange-glow);
        }

        .spinner-inner {
            position: absolute;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            width: 50px;
            height: 50px;
            border: 3px solid rgba(139, 92, 246, 0.1);
            border-bottom: 3px solid var(--accent-purple);
            border-left: 3px solid var(--accent-purple);
            border-radius: 50%;
            animation: spin 1.2s linear infinite reverse;
        }

        @keyframes spin { 
            0% { transform: rotate(0deg); } 
            100% { transform: rotate(360deg); } 
        }

        .checkmark-container { display: none; text-align: center; }

        .checkmark-circle {
            width: 100px;
            height: 100px; 
            position: relative; 
            display: inline-block;
            border-radius: 50%; 
            border: 3px solid var(--accent-green);
            margin-bottom: 24px;
            animation: success-anim 0.6s cubic-bezier(0.4, 0, 0.2, 1) forwards;
            box-shadow: 0 0 40px rgba(16, 185, 129, 0.3);
        }

        .checkmark-circle::before {
            content: '';
            display: block; 
            width: 30px; 
            height: 50px;
            border: solid var(--accent-green); 
            border-width: 0 4px 4px 0;
            position: absolute; 
            top: 15px; 
            left: 35px;
            transform: rotate(45deg); 
            opacity: 0;
            animation: checkmark-anim 0.4s 0.4s cubic-bezier(0.4, 0, 0.2, 1) forwards;
        }

        .fail-container { display: none; text-align: center; }

        .fail-circle {
            width: 100px;
            height: 100px; 
            position: relative; 
            display: inline-block;
            border-radius: 50%; 
            border: 3px solid var(--accent-red);
            margin-bottom: 24px;
            animation: fail-anim 0.6s cubic-bezier(0.4, 0, 0.2, 1) forwards;
            box-shadow: 0 0 40px rgba(239, 68, 68, 0.3);
        }

        .fail-circle::before, .fail-circle::after {
            content: '';
            position: absolute; 
            width: 4px; 
            height: 50px; 
            background: var(--accent-red);
            top: 23px; 
            left: 46px; 
            border-radius: 4px;
            animation: crossAnim 0.3s 0.4s ease-out forwards;
            opacity: 0;
        }

        .fail-circle::before { transform: rotate(45deg); }
        .fail-circle::after { transform: rotate(-45deg); }

        @keyframes success-anim { 
            0% { transform: scale(0); opacity: 0; } 
            50% { transform: scale(1.2); } 
            100% { transform: scale(1); opacity: 1; } 
        }

        @keyframes checkmark-anim { 
            0% { opacity: 0; height: 0; } 
            100% { opacity: 1; height: 50px; } 
        }

        @keyframes fail-anim { 
            0% { transform: scale(0); opacity: 0; } 
            50% { transform: scale(1.2); } 
            100% { transform: scale(1); opacity: 1; } 
        }

        @keyframes crossAnim {
            0% { opacity: 0; }
            100% { opacity: 1; }
        }

        .anim-text { 
            font-size: 24px; 
            font-weight: 800; 
            color: white;
            margin-top: 10px; 
        }

        .loading-text {
            color: var(--text-secondary);
            font-size: 15px;
            margin-top: 20px;
            font-weight: 500;
            letter-spacing: 2px;
            text-transform: uppercase;
            animation: textPulse 1.5s ease-in-out infinite;
        }

        @keyframes textPulse {
            0%, 100% { opacity: 0.5; }
            50% { opacity: 1; }
        }

        /* ENHANCED Flash Messages - Professional (No Emoji) */
        #flash-container {
            position: fixed;
            top: 20px;
            left: 50%;
            transform: translateX(-50%);
            z-index: 10001;
            width: auto;
            min-width: 350px;
            max-width: 500px;
            display: flex;
            flex-direction: column;
            align-items: center;
            gap: 10px;
        }

        .flash-message {
            padding: 18px 50px 18px 30px;
            border-radius: 16px;
            font-size: 15px;
            font-weight: 600;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 14px;
            animation: flashBounceIn 0.6s cubic-bezier(0.68, -0.55, 0.265, 1.55);
            box-shadow: 0 15px 40px rgba(0,0,0,0.4);
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255,255,255,0.1);
            position: relative;
            overflow: hidden;
        }

        .flash-message::before {
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent);
            animation: flashShine 2s ease-in-out infinite;
        }

        @keyframes flashBounceIn {
            0% { 
                opacity: 0;
                transform: translateY(-50px) scale(0.8);
            }
            50% {
                transform: translateY(10px) scale(1.05);
            }
            100% { 
                opacity: 1;
                transform: translateY(0) scale(1);
            }
        }

        @keyframes flashShine {
            0% { left: -100%; }
            50%, 100% { left: 100%; }
        }

        .flash-message i {
            font-size: 22px;
            animation: flashIconPulse 1s ease-in-out infinite;
        }

        @keyframes flashIconPulse {
            0%, 100% { transform: scale(1); }
            50% { transform: scale(1.2); }
        }

        .flash-error {
            background: linear-gradient(135deg, rgba(239, 68, 68, 0.95) 0%, rgba(185, 28, 28, 0.95) 100%);
            border-left: 4px solid #fca5a5;
            color: white;
        }

        .flash-success {
            background: linear-gradient(135deg, rgba(16, 185, 129, 0.95) 0%, rgba(5, 150, 105, 0.95) 100%);
            border-left: 4px solid #6ee7b7;
            color: white;
        }

        .flash-info {
            background: linear-gradient(135deg, rgba(59, 130, 246, 0.95) 0%, rgba(37, 99, 235, 0.95) 100%);
            border-left: 4px solid #93c5fd;
            color: white;
        }

        .flash-warning {
            background: linear-gradient(135deg, rgba(245, 158, 11, 0.95) 0%, rgba(217, 119, 6, 0.95) 100%);
            border-left: 4px solid #fcd34d;
            color: white;
        }

        .flash-close {
            position: absolute;
            right: 12px;
            top: 50%;
            transform: translateY(-50%);
            background: rgba(255,255,255,0.2);
            border: none;
            color: white;
            width: 24px;
            height: 24px;
            border-radius: 50%;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 12px;
            transition: var(--transition-smooth);
            opacity: 0.7;
        }

        .flash-close:hover {
            opacity: 1;
            background: rgba(255,255,255,0.3);
            transform: translateY(-50%) scale(1.1);
        }
        
        /* File Upload Zone */
        .upload-zone {
            border: 2px dashed var(--border-color);
            padding: 50px;
            text-align: center;
            border-radius: 16px;
            transition: var(--transition-smooth);
            cursor: pointer;
            position: relative;
            overflow: hidden;
        }

        .upload-zone:hover {
            border-color: var(--accent-orange);
            background: rgba(255, 122, 0, 0.05);
        }

        .upload-zone.dragover {
            border-color: var(--accent-orange);
            background: rgba(255, 122, 0, 0.1);
            transform: scale(1.02);
        }

        .upload-icon {
            font-size: 60px;
            color: var(--accent-orange);
            margin-bottom: 20px;
            display: inline-block;
            animation: uploadFloat 3s ease-in-out infinite;
        }

        @keyframes uploadFloat {
            0%, 100% { transform: translateY(0); }
            50% { transform: translateY(-10px); }
        }

        .upload-text {
            color: var(--accent-orange);
            font-weight: 600;
            font-size: 16px;
            margin-bottom: 8px;
        }

        .upload-hint {
            color: var(--text-secondary);
            font-size: 13px;
        }

        #file-count {
            margin-top: 20px;
            font-size: 14px;
            color: var(--accent-green);
            font-weight: 600;
        }

        /* Search Box for History */
        .search-box {
            position: relative;
            margin-bottom: 20px;
        }

        .search-box input {
            padding-left: 45px;
            background: rgba(255, 255, 255, 0.03);
        }

        .search-box i {
            position: absolute;
            left: 16px;
            top: 50%;
            transform: translateY(-50%);
            color: var(--text-secondary);
            font-size: 16px;
        }

        .search-box input:focus + i {
            color: var(--accent-orange);
        }

        /* History Filter Tabs */
        .history-tabs {
            display: flex;
            gap: 10px;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }

        .history-tab {
            padding: 8px 16px;
            background: rgba(255,255,255,0.03);
            border: 1px solid var(--border-color);
            border-radius: 20px;
            color: var(--text-secondary);
            font-size: 12px;
            font-weight: 600;
            cursor: pointer;
            transition: var(--transition-smooth);
        }

        .history-tab:hover, .history-tab.active {
            background: var(--accent-orange);
            border-color: var(--accent-orange);
            color: white;
        }

        .history-tab .tab-count {
            background: rgba(255,255,255,0.2);
            padding: 2px 8px;
            border-radius: 10px;
            margin-left: 6px;
            font-size: 10px;
        }

        /* Time Badge */
        .time-badge {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            background: rgba(255, 255, 255, 0.03);
            padding: 8px 14px;
            border-radius: 8px;
            font-size: 13px;
            color: var(--text-secondary);
        }

        .time-badge i {
            color: var(--accent-orange);
        }
        
        /* History Dropdown */
        .history-section {
            margin-top: 25px;
            padding-top: 25px;
            border-top: 1px solid var(--border-color);
        }
        
        .history-toggle {
            display: flex;
            align-items: center;
            justify-content: space-between;
            cursor: pointer;
            padding: 15px 20px;
            background: linear-gradient(145deg, rgba(139, 92, 246, 0.1), rgba(139, 92, 246, 0.02));
            border: 1px solid rgba(139, 92, 246, 0.2);
            border-radius: 14px;
            transition: var(--transition-smooth);
        }
        
        .history-toggle:hover {
            border-color: var(--accent-purple);
            transform: translateY(-2px);
        }
        
        .history-toggle-left {
            display: flex;
            align-items: center;
            gap: 12px;
        }
        
        .history-toggle-icon {
            width: 42px;
            height: 42px;
            background: linear-gradient(145deg, rgba(139, 92, 246, 0.3), rgba(139, 92, 246, 0.1));
            border-radius: 10px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 18px;
            color: var(--accent-purple);
        }
        
        .history-toggle-text {
            font-size: 15px;
            font-weight: 600;
            color: white;
        }
        
        .history-toggle-sub {
            font-size: 12px;
            color: var(--text-secondary);
            margin-top: 2px;
        }
        
        .history-badge {
            background: var(--accent-purple);
            color: white;
            padding: 5px 14px;
            border-radius: 20px;
            font-size: 13px;
            font-weight: 700;
        }
        
        .history-dropdown {
            display: none;
            margin-top: 15px;
            max-height: 320px;
            overflow-y: auto;
            padding: 5px;
        }
        
        .history-dropdown.active {
            display: block;
            animation: slideDown 0.3s ease-out;
        }
        
        @keyframes slideDown {
            from { opacity: 0; transform: translateY(-10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .history-booking-item {
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 14px 16px;
            background: rgba(255, 255, 255, 0.02);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            margin-bottom: 10px;
            cursor: pointer;
            transition: var(--transition-smooth);
            text-decoration: none;
        }
        
        .history-booking-item:hover {
            background: rgba(139, 92, 246, 0.1);
            border-color: rgba(139, 92, 246, 0.3);
            transform: translateX(5px);
        }
        
        .booking-item-left {
            display: flex;
            flex-direction: column;
            gap: 4px;
        }
        
        .booking-ref {
            font-size: 15px;
            font-weight: 700;
            color: var(--accent-purple);
        }
        
        .booking-info {
            font-size: 12px;
            color: var(--text-secondary);
        }
        
        .booking-stats {
            display: flex;
            align-items: center;
            gap: 15px;
        }
        
        .booking-stat {
            text-align: center;
        }
        
        .booking-stat-value {
            font-size: 16px;
            font-weight: 800;
            color: var(--accent-green);
        }
        
        .booking-stat-label {
            font-size: 10px;
            color: var(--text-secondary);
            text-transform: uppercase;
        }
        
        .booking-arrow {
            color: var(--text-secondary);
            font-size: 14px;
            transition: var(--transition-smooth);
        }
        
        .history-booking-item:hover .booking-arrow {
            color: var(--accent-purple);
            transform: translateX(5px);
        }
        
        .empty-history {
            text-align: center;
            padding: 30px;
            color: var(--text-secondary);
        }
        
        .empty-history i {
            font-size: 40px;
            opacity: 0.3;
            margin-bottom: 10px;
            display: block;
        }

        .no-results {
            text-align: center;
            padding: 40px 20px;
            color: var(--text-secondary);
        }

        .no-results i {
            font-size: 50px;
            opacity: 0.2;
            margin-bottom: 15px;
            display: block;
        }

        .no-results h4 {
            color: white;
            margin-bottom: 8px;
        }
        
        /* Permission Checkbox */
        .perm-checkbox {
            background: rgba(255, 255, 255, 0.03);
            padding: 14px 18px;
            border-radius: 12px;
            cursor: pointer;
            display: flex;
            align-items: center;
            border: 1px solid var(--border-color);
            transition: var(--transition-smooth);
            flex: 1;
            min-width: 100px;
        }

        .perm-checkbox:hover {
            border-color: var(--accent-orange);
            background: rgba(255, 122, 0, 0.05);
        }

        .perm-checkbox input {
            width: auto;
            margin-right: 10px;
            accent-color: var(--accent-orange);
        }

        .perm-checkbox span {
            font-size: 13px;
            font-weight: 500;
            color: var(--text-secondary);
        }

        .perm-checkbox:has(input:checked) {
            border-color: var(--accent-orange);
            background: rgba(255, 122, 0, 0.1);
        }

        .perm-checkbox:has(input:checked) span {
            color: var(--accent-orange);
        }

        /* Chart Container */
        .chart-container {
            position: relative;
            height: 280px;
            padding: 10px;
        }

        /* Realtime Indicator */
        .realtime-indicator {
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 12px;
            color: var(--text-secondary);
            padding: 6px 12px;
            background: rgba(16, 185, 129, 0.1);
            border-radius: 20px;
            border: 1px solid rgba(16, 185, 129, 0.2);
        }

        .realtime-dot {
            width: 8px;
            height: 8px;
            background: var(--accent-green);
            border-radius: 50%;
            animation: realtimePulse 1s infinite;
        }

        @keyframes realtimePulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.5; }
        }

        /* Mobile */
        .mobile-toggle { 
            display: none; 
            position: fixed; 
            top: 20px;
            right: 20px; 
            z-index: 2000; 
            color: white; 
            background: var(--bg-card);
            padding: 12px 14px; 
            border-radius: 12px;
            border: 1px solid var(--border-color);
            cursor: pointer;
        }

        @media (max-width: 1024px) {
            .sidebar { transform: translateX(-100%); } 
            .sidebar.active { transform: translateX(0); }
            .main-content { margin-left: 0; width: 100%; padding: 20px; }
            .dashboard-grid-2 { grid-template-columns: 1fr; }
            .mobile-toggle { display: flex; align-items: center; justify-content: center; }
            .header-section { flex-direction: column; gap: 15px; }
        }

        @media (max-width: 768px) {
            .stats-grid { grid-template-columns: 1fr; }
            .page-title { font-size: 24px; }
        }

        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(20px); }
            to { opacity: 1; transform: translateY(0); }
        }
    </style>
"""
# ==============================================================================
# HELPER FUNCTIONS: MongoDB Operations
# ==============================================================================

def load_users():
    record = users_col.find_one({"_id": "global_users"})
    default_users = {
        "Admin": {
            "password": "@Nijhum@12", 
            "role": "admin", 
            "permissions": ["closing", "po_sheet", "user_manage", "view_history", "accessories"],
            "created_at": "N/A",
            "last_login": "Never",
            "last_duration": "N/A"
        }
    }
    if record:
        return record['data']
    else:
        users_col.insert_one({"_id": "global_users", "data": default_users})
        return default_users

def save_users(users_data):
    users_col.replace_one(
        {"_id": "global_users"}, 
        {"_id": "global_users", "data": users_data}, 
        upsert=True
    )

def load_stats():
    record = stats_col.find_one({"_id": "dashboard_stats"})
    if record:
        return record['data']
    else:
        default_stats = {"downloads": [], "last_booking": "None"}
        stats_col.insert_one({"_id": "dashboard_stats", "data": default_stats})
        return default_stats

def save_stats(data):
    stats_col.replace_one(
        {"_id": "dashboard_stats"},
        {"_id": "dashboard_stats", "data": data},
        upsert=True
    )

def update_stats(ref_no, username):
    data = load_stats()
    now = get_bd_time()
    new_record = {
        "ref": ref_no,
        "user": username,
        "date": now.strftime('%d-%m-%Y'),
        "time": now.strftime('%I:%M %p'),
        "type": "Closing Report",
        "iso_time": now.isoformat()
    }
    data['downloads'].insert(0, new_record)
    if len(data['downloads']) > 5000:
        data['downloads'] = data['downloads'][:5000]
    data['last_booking'] = ref_no
    save_stats(data)

def update_po_stats(username, file_count, booking_ref="N/A"):
    data = load_stats()
    now = get_bd_time()
    new_record = {
        "ref": booking_ref,
        "user": username,
        "file_count": file_count,
        "date": now.strftime('%d-%m-%Y'),
        "time": now.strftime('%I:%M %p'),
        "type": "PO Sheet",
        "iso_time": now.isoformat()
    }
    if 'downloads' not in data:
        data['downloads'] = []
    data['downloads'].insert(0, new_record)
    if len(data['downloads']) > 5000:
        data['downloads'] = data['downloads'][:5000]
    save_stats(data)

def update_accessories_stats(ref_no, username, challan_info=""):
    data = load_stats()
    now = get_bd_time()
    new_record = {
        "ref": ref_no,
        "user": username,
        "date": now.strftime('%d-%m-%Y'),
        "time": now.strftime('%I:%M %p'),
        "type": "Accessories",
        "info": challan_info,
        "iso_time": now.isoformat()
    }
    if 'downloads' not in data:
        data['downloads'] = []
    data['downloads'].insert(0, new_record)
    if len(data['downloads']) > 5000:
        data['downloads'] = data['downloads'][:5000]
    save_stats(data)

def load_accessories_db():
    record = accessories_col.find_one({"_id": "accessories_data"})
    if record:
        return record['data']
    else:
        return {}

def save_accessories_db(data):
    accessories_col.replace_one(
        {"_id": "accessories_data"},
        {"_id": "accessories_data", "data": data},
        upsert=True
    )

def get_all_accessories_bookings():
    db_acc = load_accessories_db()
    bookings = []
    for ref, data in db_acc.items():
        challan_count = len(data.get('challans', []))
        total_qty = sum([int(c.get('qty', 0)) for c in data.get('challans', [])])
        bookings.append({
            'ref': ref,
            'buyer': data.get('buyer', 'N/A'),
            'style': data.get('style', 'N/A'),
            'challan_count': challan_count,
            'total_qty': total_qty,
            'last_updated': data.get('last_api_call', 'N/A')
        })
    bookings = sorted(bookings, key=lambda x: x['ref'], reverse=True)
    return bookings

# ==============================================================================
# 24 HOUR HISTORY FUNCTIONS
# ==============================================================================

def get_24h_history():
    stats_data = load_stats()
    all_history = stats_data.get('downloads', [])
    
    now = get_bd_time()
    cutoff_time = now - timedelta(hours=24)
    
    filtered_history = []
    for item in all_history:
        try:
            if 'iso_time' in item:
                item_time = datetime.fromisoformat(item['iso_time'])
                if item_time.tzinfo is None:
                    item_time = bd_tz.localize(item_time)
                if item_time >= cutoff_time:
                    filtered_history.append(item)
            else:
                date_str = item.get('date', '')
                time_str = item.get('time', '')
                if date_str and time_str:
                    dt_str = f"{date_str} {time_str}"
                    try:
                        item_time = datetime.strptime(dt_str, '%d-%m-%Y %I:%M %p')
                        item_time = bd_tz.localize(item_time)
                        if item_time >= cutoff_time:
                            filtered_history.append(item)
                    except:
                        pass
        except:
            pass
    
    return filtered_history

def search_history(query, history_list=None):
    if history_list is None:
        stats_data = load_stats()
        history_list = stats_data.get('downloads', [])
    
    if not query or query.strip() == '':
        return history_list
    
    query = query.strip().lower()
    results = []
    
    for item in history_list:
        searchable = f"{item.get('ref', '')} {item.get('user', '')} {item.get('type', '')} {item.get('date', '')} {item.get('time', '')} {item.get('info', '')}".lower()
        if query in searchable:
            results.append(item)
    
    return results

# ==============================================================================
# DASHBOARD SUMMARY FUNCTION (Updated with 24h History)
# ==============================================================================

def get_dashboard_summary_v2():
    stats_data = load_stats()
    acc_db = load_accessories_db()
    users_data = load_users()
    
    now = get_bd_time()
    today_str = now.strftime('%d-%m-%Y')
    
    user_details = []
    for u, d in users_data.items():
        user_details.append({
            "username": u,
            "role": d.get('role', 'user'),
            "created_at": d.get('created_at', 'N/A'),
            "last_login": d.get('last_login', 'Never'),
            "last_duration": d.get('last_duration', 'N/A')
        })

    acc_lifetime_count = 0
    acc_today_list = []
    daily_data = defaultdict(lambda: {'closing': 0, 'po': 0, 'acc': 0})

    for ref, data in acc_db.items():
        for challan in data.get('challans', []):
            acc_lifetime_count += 1
            c_date = challan.get('date')
            if c_date == today_str:
                acc_today_list.append({
                    "ref": ref,
                    "buyer": data.get('buyer'),
                    "style": data.get('style'),
                    "time": "Today",
                    "qty": challan.get('qty')
                })
            
            try:
                dt_obj = datetime.strptime(c_date, '%d-%m-%Y')
                sort_key = dt_obj.strftime('%Y-%m-%d')
                daily_data[sort_key]['acc'] += 1
                daily_data[sort_key]['label'] = dt_obj.strftime('%d-%b')
            except: 
                pass

    closing_lifetime_count = 0
    po_lifetime_count = 0
    closing_list = []
    po_list = []
    
    history = stats_data.get('downloads', [])
    for item in history:
        item_date = item.get('date', '')
        if item.get('type') == 'PO Sheet':
            po_lifetime_count += 1
            if item_date == today_str: 
                po_list.append(item)
        elif item.get('type') == 'Accessories':
            pass
        else:
            closing_lifetime_count += 1
            if item_date == today_str:
                closing_list.append(item)
        
        try:
            dt_obj = datetime.strptime(item_date, '%d-%m-%Y')
            sort_key = dt_obj.strftime('%Y-%m-%d')
            
            if item.get('type') == 'PO Sheet':
                daily_data[sort_key]['po'] += 1
            elif item.get('type') == 'Accessories':
                daily_data[sort_key]['acc'] += 1
            else: 
                daily_data[sort_key]['closing'] += 1
            daily_data[sort_key]['label'] = dt_obj.strftime('%d-%b')
        except:
            pass
    
    first_of_last_month = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    start_date = first_of_last_month.strftime('%Y-%m-%d')
    end_date = now.strftime('%Y-%m-%d')
    
    sorted_keys = sorted([k for k in daily_data.keys() if start_date <= k <= end_date])
    
    chart_labels = []
    chart_closing = []
    chart_po = []
    chart_acc = []

    if not sorted_keys:
        curr_d = now.strftime('%d-%b')
        chart_labels = [curr_d]
        chart_closing = [0]
        chart_po = [0]
        chart_acc = [0]
    else:
        for k in sorted_keys:
            d = daily_data[k]
            chart_labels.append(d.get('label', k))
            chart_closing.append(d['closing'])
            chart_po.append(d['po'])
            chart_acc.append(d['acc'])

    history_24h = get_24h_history()
    
    closing_24h = len([h for h in history_24h if h.get('type') == 'Closing Report'])
    po_24h = len([h for h in history_24h if h.get('type') == 'PO Sheet'])
    acc_24h = len([h for h in history_24h if h.get('type') == 'Accessories'])

    return {
        "users": {"count": len(users_data), "details": user_details},
        "accessories": {"count": acc_lifetime_count, "details": acc_today_list},
        "closing": {"count": closing_lifetime_count, "details": closing_list},
        "po": {"count": po_lifetime_count, "details": po_list},
        "chart": {
            "labels": chart_labels,
            "closing": chart_closing,
            "po": chart_po,
            "acc": chart_acc
        },
        "history": history,
        "history_24h": history_24h,
        "counts_24h": {
            "closing": closing_24h,
            "po": po_24h,
            "acc": acc_24h,
            "total": len(history_24h)
        }
    }

# ==============================================================================
# NEW PO SHEET PARSER LOGIC (আপনার দেওয়া কোড থেকে হুবহু)
# ==============================================================================

def is_potential_size(header):
    h = header.strip().upper()
    if h in ["COLO", "SIZE", "TOTAL", "QUANTITY", "PRICE", "AMOUNT", "CURRENCY", "ORDER NO", "P.O NO"]:
        return False
    if re.match(r'^\d+$', h): return True
    if re.match(r'^\d+[AMYT]$', h): return True
    if re.match(r'^(XXS|XS|S|M|L|XL|XXL|XXXL|TU|ONE\s*SIZE)$', h): return True
    if re.match(r'^[A-Z]\d{2,}$', h): return False
    return False

def sort_sizes(size_list):
    STANDARD_ORDER = [
        '0M', '1M', '3M', '6M', '9M', '12M', '18M', '24M', '36M',
        '2A', '3A', '4A', '5A', '6A', '8A', '10A', '12A', '14A', '16A', '18A',
        'XXS', 'XS', 'S', 'M', 'L', 'XL', 'XXL', '3XL', '4XL', '5XL',
        'TU', 'One Size'
    ]
    def sort_key(s):
        s = s.strip()
        if s in STANDARD_ORDER: return (0, STANDARD_ORDER.index(s))
        if s.isdigit(): return (1, int(s))
        match = re.match(r'^(\d+)([A-Z]+)$', s)
        if match: return (2, int(match.group(1)), match.group(2))
        return (3, s)
    return sorted(size_list, key=sort_key)

def extract_metadata(first_page_text):
    meta = {
        'buyer': 'N/A', 'booking': 'N/A', 'style': 'N/A', 
        'season': 'N/A', 'dept': 'N/A', 'item': 'N/A'
    }
    
    if "KIABI" in first_page_text.upper():
        meta['buyer'] = "KIABI"
    else:
        buyer_match = re.search(r"Buyer.*?Name[\s\S]*?([\w\s&]+)(?:\n|$)", first_page_text)
        if buyer_match: meta['buyer'] = buyer_match.group(1).strip()

    booking_block_match = re.search(r"(?:Internal )?Booking NO\.?[:\s]*([\s\S]*?)(?:System NO|Control No|Buyer)", first_page_text, re.IGNORECASE)
    if booking_block_match: 
        raw_booking = booking_block_match.group(1).strip()
        clean_booking = raw_booking.replace('\n', '').replace('\r', '').replace(' ', '')
        if "System" in clean_booking: clean_booking = clean_booking.split("System")[0]
        meta['booking'] = clean_booking

    style_match = re.search(r"Style Ref\.?[:\s]*([\w-]+)", first_page_text, re.IGNORECASE)
    if style_match: meta['style'] = style_match.group(1).strip()
    else:
        style_match = re.search(r"Style Des\.?[\s\S]*?([\w-]+)", first_page_text, re.IGNORECASE)
        if style_match: meta['style'] = style_match.group(1).strip()

    season_match = re.search(r"Season\s*[:\n\"]*([\w\d-]+)", first_page_text, re.IGNORECASE)
    if season_match: meta['season'] = season_match.group(1).strip()

    dept_match = re.search(r"Dept\.?[\s\n:]*([A-Za-z]+)", first_page_text, re.IGNORECASE)
    if dept_match: meta['dept'] = dept_match.group(1).strip()

    item_match = re.search(r"Garments? Item[\s\n:]*([^\n\r]+)", first_page_text, re.IGNORECASE)
    if item_match: 
        item_text = item_match.group(1).strip()
        if "Style" in item_text: item_text = item_text.split("Style")[0].strip()
        meta['item'] = item_text

    return meta

def is_color_name(text):
    text = text.strip()
    if not text:
        return False
    if re.match(r'^\d+$', text):
        return False
    if re.match(r'^\d+[,\.]\d{2}$', text):
        return False
    if is_potential_size(text):
        return False
    keywords = ['spec', 'price', 'total', 'quantity', 'amount', 'currency', 'order']
    if any(kw in text.lower() for kw in keywords):
        return False
    if not re.search(r'[a-zA-Z]', text):
        return False
    return True

def is_partial_color_name(text):
    text = text.strip()
    if not text:
        return False
    if re.match(r'^[A-Za-z\s]+$', text):
        keywords = ['spec', 'price', 'total', 'quantity', 'amount']
        if not any(kw in text.lower() for kw in keywords):
            return True
    return False

def parse_vertical_table(lines, start_idx, sizes, order_no):
    extracted_data = []
    i = start_idx
    
    while i < len(lines):
        line = lines[i].strip()
        
        if line.startswith("Total") and i + 1 < len(lines):
            next_line = lines[i + 1].strip() if i + 1 < len(lines) else ""
            if "Quantity" in next_line or "Amount" in next_line or re.match(r'^Quantity', next_line):
                break
            if re.match(r'^\d', next_line):
                break
        
        if line and is_color_name(line):
            color_name = line
            i += 1
            
            while i < len(lines):
                next_line = lines[i].strip()
                
                if 'spec' in next_line.lower():
                    i += 1
                    break
                
                if re.match(r'^\d+$', next_line):
                    break
                
                if not next_line:
                    i += 1
                    continue
                
                if is_partial_color_name(next_line):
                    color_name = color_name + " " + next_line
                    i += 1
                else:
                    break
            
            if i < len(lines) and 'spec' in lines[i].lower():
                i += 1
            
            quantities = []
            size_idx = 0
            
            while size_idx < len(sizes) and i < len(lines):
                qty_line = lines[i].strip() if i < len(lines) else ""
                price_line = lines[i + 1].strip() if i + 1 < len(lines) else ""
                
                if qty_line and is_color_name(qty_line):
                    while size_idx < len(sizes):
                        quantities.append(0)
                        size_idx += 1
                    break
                
                if (qty_line == "" or qty_line.isspace()) and (price_line == "" or price_line.isspace()):
                    quantities.append(0)
                    size_idx += 1
                    i += 2
                    continue
                
                if re.match(r'^\d+$', qty_line):
                    quantities.append(int(qty_line))
                    size_idx += 1
                    i += 2
                    continue
                
                i += 1
            
            if quantities:
                while len(quantities) < len(sizes):
                    quantities.append(0)
                
                for idx, size in enumerate(sizes):
                    extracted_data.append({
                        'P.O NO': order_no,
                        'Color': color_name,
                        'Size': size,
                        'Quantity': quantities[idx] if idx < len(quantities) else 0
                    })
            
            continue
        
        i += 1
    
    return extracted_data

def extract_data_dynamic(file_path):
    extracted_data = []
    metadata = {
        'buyer': 'N/A', 'booking': 'N/A', 'style': 'N/A', 
        'season': 'N/A', 'dept': 'N/A', 'item': 'N/A'
    }
    order_no = "Unknown"
    
    try:
        reader = pypdf.PdfReader(file_path)
        first_page_text = reader.pages[0].extract_text()
        
        if "Main Fabric Booking" in first_page_text or "Fabric Booking Sheet" in first_page_text:
            metadata = extract_metadata(first_page_text)
            return [], metadata 

        order_match = re.search(r"Order no\D*(\d+)", first_page_text, re.IGNORECASE)
        if order_match: order_no = order_match.group(1)
        else:
            alt_match = re.search(r"Order\s*[:\.]?\s*(\d+)", first_page_text, re.IGNORECASE)
            if alt_match: order_no = alt_match.group(1)
        
        order_no = str(order_no).strip()
        if order_no.endswith("00"): order_no = order_no[:-2]

        for page in reader.pages:
            text = page.extract_text()
            lines = text.split('\n')
            
            for i, line in enumerate(lines):
                if ("Colo" in line or "Size" in line) and "Total" in line:
                    parts = line.split()
                    try:
                        total_idx = [idx for idx, x in enumerate(parts) if 'Total' in x][0]
                        raw_sizes = parts[:total_idx]
                        sizes = [s for s in raw_sizes if s not in ["Colo", "/", "Size", "Colo/Size", "Colo/", "Size's"]]
                        
                        valid_size_count = sum(1 for s in sizes if is_potential_size(s))
                        if sizes and valid_size_count >= len(sizes) / 2:
                            data = parse_vertical_table(lines, i + 1, sizes, order_no)
                            extracted_data.extend(data)
                    except: 
                        pass
                    break
                    
    except Exception as e: 
        print(f"Error processing file: {e}")
    
    return extracted_data, metadata

# ==============================================================================
# CLOSING REPORT API & EXCEL GENERATION
# ==============================================================================

def get_authenticated_session(username, password):
    login_url = 'http://180.92.235.190:8022/erp/login.php'
    login_payload = {'txt_userid': username, 'txt_password': password, 'submit': 'Login'}
    session_req = requests.Session()
    session_req.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    })
    try:
        response = session_req.post(login_url, data=login_payload, timeout=300)
        if "dashboard.php" in response.url or "Invalid" not in response.text:
            return session_req
        else: 
            return None
    except requests.exceptions.RequestException as e:
        print(f"Connection Error: {e}")
        return None

def fetch_closing_report_data(internal_ref_no):
    active_session = get_authenticated_session("input2.clothing-cutting", "123456")
    if not active_session: 
        return None

    report_url = 'http://180.92.235.190:8022/erp/prod_planning/reports/requires/cutting_lay_production_report_controller.php'
    payload_template = {
        'action': 'report_generate', 
        'cbo_wo_company_name': '2', 
        'cbo_location_name': '2', 
        'cbo_floor_id': '0', 
        'cbo_buyer_name': '0', 
        'txt_internal_ref_no': internal_ref_no, 
        'report_type': 'cutting_order_to_input',
        'cbo_company_name': '2',
        'cbo_year_selection': '2025'
    }
    found_data = None
   
    for year in ['2025', '2024', '2023']: 
        for company_id in range(1, 6):
            payload = payload_template.copy()
            payload['cbo_year_selection'] = year
            payload['cbo_company_name'] = str(company_id)
            try:
                response = active_session.post(report_url, data=payload, timeout=300)
                if response.status_code == 200 and "Data not Found" not in response.text and len(response.text) > 500:
                    found_data = response.text
                    break
            except: 
                continue
        if found_data: 
            break
    
    if found_data:
        return parse_report_data(found_data)
    return None

def parse_report_data(html_content):
    all_report_data = []
    try:
        soup = BeautifulSoup(html_content, 'lxml')
        header_row = soup.select_one('thead tr:nth-of-type(2)')
        if not header_row: 
            return None
        all_th = header_row.find_all('th')
        headers = [th.get_text(strip=True) for th in all_th if 'total' not in th.get_text(strip=True).lower()]
        data_rows = soup.select('div#scroll_body table tbody tr')
        item_blocks = []
        current_block = []
        for row in data_rows:
            if row.get('bgcolor') == '#cddcdc':
                if current_block: item_blocks.append(current_block)
                current_block = []
            else:
                current_block.append(row)
        if current_block: item_blocks.append(current_block)
        
        for block in item_blocks:  
            style, color, buyer_name, gmts_qty_data, sewing_input_data, cutting_qc_data = "N/A", "N/A", "N/A", None, None, None
            for row in block:
                cells = row.find_all('td')
                if len(cells) > 2:
                    criteria_main = cells[0].get_text(strip=True)
                    criteria_sub = cells[2].get_text(strip=True)
                    main_lower, sub_lower = criteria_main.lower(), criteria_sub.lower()
                    
                    if main_lower == "style": style = cells[1].get_text(strip=True)
                    elif main_lower == "color & gmts. item": color = cells[1].get_text(strip=True)
                    elif "buyer" in main_lower: buyer_name = cells[1].get_text(strip=True)
                    
                    if sub_lower == "gmts. color /country qty": gmts_qty_data = [cell.get_text(strip=True) for cell in cells[3:len(headers)+3]]
                    
                    if "sewing input" in main_lower: sewing_input_data = [cell.get_text(strip=True) for cell in cells[1:len(headers)+1]]
                    elif "sewing input" in sub_lower: sewing_input_data = [cell.get_text(strip=True) for cell in cells[3:len(headers)+3]]
    
                    if "cutting qc" in main_lower and "balance" not in main_lower: 
                        cutting_qc_data = [cell.get_text(strip=True) for cell in cells[1:len(headers)+1]]
                    elif "cutting qc" in sub_lower and "balance" not in sub_lower: 
                        cutting_qc_data = [cell.get_text(strip=True) for cell in cells[3:len(headers)+3]]
            if gmts_qty_data:  
                plus_3_percent_data = []
                for value in gmts_qty_data:  
                    try:
                        new_qty = round(int(value.replace(',', '')) * 1.03)
                        plus_3_percent_data.append(str(new_qty))
                    except (ValueError, TypeError):
                        plus_3_percent_data.append(value)
                all_report_data.append({
                    'style': style, 'buyer': buyer_name, 'color': color, 
                    'headers': headers, 'gmts_qty': gmts_qty_data, 
                    'plus_3_percent': plus_3_percent_data, 
                    'sewing_input': sewing_input_data if sewing_input_data else [], 
                    'cutting_qc': cutting_qc_data if cutting_qc_data else []
                })
        return all_report_data
    except Exception as e: 
        print(f"Parse error: {e}")
        return None

def create_formatted_excel_report(report_data, internal_ref_no=""):
    if not report_data: 
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Closing Report"
    
    bold_font = Font(bold=True)
    title_font = Font(size=32, bold=True, color="7B261A") 
    white_bold_font = Font(size=16.5, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    color_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    medium_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    
    ir_ib_fill = PatternFill(start_color="7B261A", end_color="7B261A", fill_type="solid") 
    header_row_fill = PatternFill(start_color="DE7465", end_color="DE7465", fill_type="solid") 
    light_brown_fill = PatternFill(start_color="DE7465", end_color="DE7465", fill_type="solid") 
    light_blue_fill = PatternFill(start_color="B9C2DF", end_color="B9C2DF", fill_type="solid") 
    light_green_fill = PatternFill(start_color="C4D09D", end_color="C4D09D", fill_type="solid") 
    dark_green_fill = PatternFill(start_color="f1f2e8", end_color="f1f2e8", fill_type="solid") 

    NUM_COLUMNS, TABLE_START_ROW = 9, 8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLUMNS)
    
    ws['A1'].value = "COTTON CLOTHING BD LTD"
    ws['A1'].font = title_font 
    ws['A1'].alignment = center_align

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLUMNS)
    ws['A2'].value = "CLOSING REPORT [ INPUT SECTION ]"
    ws['A2'].font = Font(size=15, bold=True) 
    ws['A2'].alignment = center_align
    ws.row_dimensions[3].height = 6

    formatted_ref_no = internal_ref_no.upper()
    current_date = get_bd_time().strftime("%d/%m/%Y")
    
    left_sub_headers = {
        'A4': 'BUYER', 'B4': report_data[0].get('buyer', ''), 
        'A5': 'IR/IB NO', 'B5': formatted_ref_no, 
        'A6': 'STYLE NO', 'B6': report_data[0].get('style', '')
    }
    
    for cell_ref, value in left_sub_headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = bold_font
        cell.alignment = left_align
        cell.border = thin_border
        if cell_ref == 'B5':
            cell.fill = ir_ib_fill 
            cell.font = white_bold_font 
        else:
            cell.fill = dark_green_fill 

    ws.merge_cells('B4:G4'); ws.merge_cells('B5:G5'); ws.merge_cells('B6:G6')
    
    right_sub_headers = {'H4': 'CLOSING DATE', 'I4': current_date, 'H5': 'SHIPMENT', 'I5': 'ALL', 'H6': 'PO NO', 'I6': 'ALL'}
    for cell_ref, value in right_sub_headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = bold_font
        cell.alignment = left_align
        cell.border = thin_border
        cell.fill = dark_green_fill 

    for row in range(4, 7):
        for col in range(3, 8): 
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
       
    current_row = TABLE_START_ROW
    for block in report_data:
        table_headers = ["COLOUR NAME", "SIZE", "ORDER QTY 3%", "ACTUAL QTY", "CUTTING QC", "INPUT QTY", "BALANCE", "SHORT/PLUS QTY", "Percentage %"]
        for col_idx, header in enumerate(table_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = medium_border
            cell.fill = header_row_fill 

        current_row += 1
        start_merge_row = current_row
        full_color_name = block.get('color', 'N/A')

        for i, size in enumerate(block['headers']):
            color_to_write = full_color_name if i == 0 else ""
            actual_qty = int(block['gmts_qty'][i].replace(',', '') or 0)
            input_qty = int(block['sewing_input'][i].replace(',', '') or 0) if i < len(block['sewing_input']) else 0
            cutting_qc_val = int(block.get('cutting_qc', [])[i].replace(',', '') or 0) if i < len(block.get('cutting_qc', [])) else 0
            
            ws.cell(row=current_row, column=1, value=color_to_write)
            ws.cell(row=current_row, column=2, value=size)
            ws.cell(row=current_row, column=4, value=actual_qty)
            ws.cell(row=current_row, column=5, value=cutting_qc_val)
            ws.cell(row=current_row, column=6, value=input_qty)
            
            ws.cell(row=current_row, column=3, value=f"=ROUND(D{current_row}*1.03, 0)")      
            ws.cell(row=current_row, column=7, value=f"=E{current_row}-F{current_row}")      
            ws.cell(row=current_row, column=8, value=f"=F{current_row}-C{current_row}")      
            ws.cell(row=current_row, column=9, value=f'=IF(C{current_row}<>0, H{current_row}/C{current_row}, 0)') 
            
            for col_idx in range(1, NUM_COLUMNS + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = medium_border if col_idx == 2 else thin_border
                cell.alignment = center_align
    
                if col_idx in [1, 2, 3, 6, 9]: cell.font = bold_font
                
                if col_idx == 3: cell.fill = light_blue_fill      
                elif col_idx == 6: cell.fill = light_green_fill   
                else: cell.fill = dark_green_fill 

                if col_idx == 9:  
                    cell.number_format = '0.00%' 
            current_row += 1
            
        end_merge_row = current_row - 1
        if start_merge_row <= end_merge_row:
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=end_merge_row, end_column=1)
            merged_cell = ws.cell(row=start_merge_row, column=1)
            merged_cell.alignment = color_align
            if not merged_cell.font.bold: merged_cell.font = bold_font
        
        total_row_str = str(current_row)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        
        totals_formulas = {
            "A": "TOTAL",
            "C": f"=SUM(C{start_merge_row}:C{end_merge_row})",
            "D": f"=SUM(D{start_merge_row}:D{end_merge_row})",
            "E": f"=SUM(E{start_merge_row}:E{end_merge_row})",
            "F": f"=SUM(F{start_merge_row}:F{end_merge_row})",
            "G": f"=SUM(G{start_merge_row}:G{end_merge_row})",
            "H": f"=SUM(H{start_merge_row}:H{end_merge_row})",
            "I": f"=IF(C{total_row_str}<>0, H{total_row_str}/C{total_row_str}, 0)"
        }
        
        for col_letter, value_or_formula in totals_formulas.items():
            cell = ws[f"{col_letter}{current_row}"]
            cell.value = value_or_formula
            cell.font = bold_font
            cell.border = medium_border
            cell.alignment = center_align
            cell.fill = light_brown_fill 
            if col_letter == 'I':  
                cell.number_format = '0.00%'
        
        for col_idx in range(2, NUM_COLUMNS + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            if not cell.value:  
                cell.fill = dark_green_fill 
                cell.border = medium_border
        current_row += 2
    
    image_row = current_row + 1
   
    try:
        direct_image_url = 'https://i.ibb.co/v6bp0jQW/rockybilly-regular.webp'
        image_response = requests.get(direct_image_url, timeout=10)
        image_response.raise_for_status()
        original_img = PILImage.open(BytesIO(image_response.content))
        padded_img = PILImage.new('RGBA', (original_img.width + 400, original_img.height), (0, 0, 0, 0))
        padded_img.paste(original_img, (400, 0))
        padded_image_io = BytesIO()
        padded_img.save(padded_image_io, format='PNG')
        img = Image(padded_image_io)
        aspect_ratio = padded_img.height / padded_img.width
        img.width = 95
        img.height = int(img.width * aspect_ratio)
        ws.row_dimensions[image_row].height = img.height * 0.90
        ws.add_image(img, f'A{image_row}')
    except Exception:  
        pass

    signature_row = image_row + 1
    ws.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row, end_column=NUM_COLUMNS)
    titles = ["Prepared By", "Input Incharge", "Cutting Incharge", "IE & Planning", "Sewing Manager", "Cutting Manager"]
    signature_cell = ws.cell(row=signature_row, column=1)
    signature_cell.value = "                 ".join(titles)
    signature_cell.font = Font(bold=True, size=15)
    signature_cell.alignment = Alignment(horizontal='center', vertical='center')

    last_data_row = current_row - 2
    for row in ws.iter_rows(min_row=4, max_row=last_data_row):
        for cell in row:  
            if cell.coordinate == 'B5': continue
            if cell.font:  
                existing_font = cell.font
                if cell.row != 1:  
                    new_font = Font(
                        name=existing_font.name, 
                        size=16.5, 
                        bold=existing_font.bold, 
                        italic=existing_font.italic, 
                        vertAlign=existing_font.vertAlign, 
                        underline=existing_font.underline, 
                        color=existing_font.color
                    )
                    cell.font = new_font
    
    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['B'].width = 8.5
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 13.5
    ws.column_dimensions['H'].width = 23
    ws.column_dimensions['I'].width = 18
   
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0 
    ws.page_setup.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
   
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream
    # ==============================================================================
# LOGIN TEMPLATE
# ==============================================================================

LOGIN_TEMPLATE = f"""
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>Login - MNM Software</title>
    {COMMON_STYLES}
    <style>
        html, body {{
            height: 100%;
            margin: 0;
            padding: 0;
            overflow-x: hidden;
        }}
        
        body {{
            background: var(--bg-body);
            min-height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            position: relative;
            overflow-y: auto;
        }}
        
        .bg-orb {{
            position: fixed;
            border-radius: 50%;
            filter: blur(80px);
            opacity: 0.4;
            animation: orbFloat 20s ease-in-out infinite;
            pointer-events: none;
        }}
        
        .orb-1 {{
            width: 300px;
            height: 300px;
            background: var(--accent-orange);
            top: -100px;
            left: -100px;
        }}
        
        .orb-2 {{
            width: 250px;
            height: 250px;
            background: var(--accent-purple);
            bottom: -50px;
            right: -50px;
            animation-delay: -5s;
        }}
        
        .orb-3 {{
            width: 150px;
            height: 150px;
            background: var(--accent-green);
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            animation-delay: -10s;
        }}
        
        @keyframes orbFloat {{
            0%, 100% {{ transform: translate(0, 0) scale(1); }}
            25% {{ transform: translate(30px, -30px) scale(1.05); }}
            50% {{ transform: translate(-20px, 20px) scale(0.95); }}
            75% {{ transform: translate(15px, 30px) scale(1.02); }}
        }}
        
        .login-container {{
            position: relative;
            z-index: 10;
            width: 100%;
            max-width: 420px;
            padding: 20px;
            margin: auto;
            display: flex;
            flex-direction: column;
            justify-content: center;
            min-height: 100vh;
        }}
        
        .login-card {{
            background: var(--gradient-card);
            border: 1px solid var(--border-color);
            border-radius: 24px;
            padding: 40px 35px;
            backdrop-filter: blur(20px);
            box-shadow: 0 25px 80px rgba(0, 0, 0, 0.5), 0 0 60px var(--accent-orange-glow);
            animation: loginCardAppear 0.8s cubic-bezier(0.4, 0, 0.2, 1);
        }}
        
        @keyframes loginCardAppear {{
            from {{ opacity: 0; transform: translateY(30px) scale(0.95); }}
            to {{ opacity: 1; transform: translateY(0) scale(1); }}
        }}
        
        .brand-section {{
            text-align: center;
            margin-bottom: 35px;
        }}
        
        .brand-icon {{
            width: 70px;
            height: 70px;
            background: var(--gradient-orange);
            border-radius: 18px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            font-size: 32px;
            color: white;
            margin-bottom: 18px;
            box-shadow: 0 15px 40px var(--accent-orange-glow);
            animation: brandIconPulse 3s ease-in-out infinite;
        }}
        
        @keyframes brandIconPulse {{
            0%, 100% {{ transform: scale(1) rotate(0deg); }}
            50% {{ transform: scale(1.05) rotate(5deg); }}
        }}
        
        .brand-name {{
            font-size: 28px;
            font-weight: 900;
            color: white;
            letter-spacing: -1px;
        }}
        
        .brand-name span {{
            background: var(--gradient-orange);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }}
        
        .brand-tagline {{
            color: var(--text-secondary);
            font-size: 11px;
            letter-spacing: 2px;
            margin-top: 6px;
            font-weight: 600;
            text-transform: uppercase;
        }}
        
        .login-form .input-group {{ margin-bottom: 20px; }}
        
        .login-form .input-group label {{
            display: flex;
            align-items: center;
            gap: 8px;
            margin-bottom: 8px;
        }}
        
        .login-form .input-group label i {{
            color: var(--accent-orange);
            font-size: 13px;
        }}
        
        .login-form input {{
            padding: 14px 18px;
            font-size: 14px;
            border-radius: 12px;
        }}
        
        .login-btn {{
            margin-top: 8px;
            padding: 14px 24px;
            font-size: 15px;
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 10px;
        }}
        
        .login-btn i {{ transition: transform 0.3s; }}
        .login-btn:hover i {{ transform: translateX(5px); }}
        
        .footer-credit {{
            text-align: center;
            margin-top: 25px;
            color: var(--text-secondary);
            font-size: 11px;
            opacity: 0.5;
            font-weight: 500;
        }}
        
        .footer-credit a {{
            color: var(--accent-orange);
            text-decoration: none;
        }}
        
        @media (max-width: 480px) {{
            .login-container {{ padding: 15px; }}
            .login-card {{ padding: 30px 25px; border-radius: 20px; }}
            .brand-icon {{ width: 60px; height: 60px; font-size: 28px; }}
            .brand-name {{ font-size: 24px; }}
        }}
    </style>
</head>
<body>
    <div class="bg-orb orb-1"></div>
    <div class="bg-orb orb-2"></div>
    <div class="bg-orb orb-3"></div>
    
    <div id="flash-container">
        {{% with messages = get_flashed_messages(with_categories=true) %}}
            {{% if messages %}}
                {{% for category, message in messages %}}
                    <div class="flash-message flash-{{{{ category if category != 'message' else 'info' }}}}">
                        {{% if category == 'success' %}}
                            <i class="fas fa-check-circle"></i>
                        {{% elif category == 'error' %}}
                            <i class="fas fa-times-circle"></i>
                        {{% elif category == 'warning' %}}
                            <i class="fas fa-exclamation-triangle"></i>
                        {{% else %}}
                            <i class="fas fa-info-circle"></i>
                        {{% endif %}}
                        <span>{{{{ message }}}}</span>
                        <button class="flash-close" onclick="this.parentElement.remove()">
                            <i class="fas fa-times"></i>
                        </button>
                    </div>
                {{% endfor %}}
            {{% endif %}}
        {{% endwith %}}
    </div>

    <div class="login-container">
        <div class="login-card">
            <div class="brand-section">
                <div class="brand-icon">
                    <i class="fas fa-layer-group"></i>
                </div>
                <div class="brand-name">MNM<span>Software</span></div>
                <div class="brand-tagline">Secure Access Portal</div>
            </div>
            
            <form action="/login" method="post" class="login-form">
                <div class="input-group">
                    <label><i class="fas fa-user"></i> USERNAME</label>
                    <input type="text" name="username" required placeholder="Enter your ID" autocomplete="off">
                </div>
                <div class="input-group">
                    <label><i class="fas fa-lock"></i> PASSWORD</label>
                    <input type="password" name="password" required placeholder="Enter your Password">
                </div>
                <button type="submit" class="login-btn">
                    Sign In <i class="fas fa-arrow-right"></i>
                </button>
            </form>
            
            <div class="footer-credit">
                © 2025 <a href="#">Mehedi Hasan</a> - All Rights Reserved
            </div>
        </div>
    </div>
    
    <script>
        setTimeout(function() {{
            let flashMessages = document.querySelectorAll('.flash-message');
            flashMessages.forEach(function(msg) {{
                msg.style.animation = 'flashSlideOut 0.5s ease-out forwards';
                setTimeout(function() {{ msg.remove(); }}, 500);
            }});
        }}, 5000);

        const style = document.createElement('style');
        style.textContent = `
            @keyframes flashSlideOut {{
                from {{ opacity: 1; transform: translateY(0); }}
                to {{ opacity: 0; transform: translateY(-30px); }}
            }}
        `;
        document.head.appendChild(style);
    </script>
</body>
</html>
"""

# ==============================================================================
# ADMIN DASHBOARD TEMPLATE (Updated with 24h History, Search, Accessories)
# ==============================================================================

ADMIN_DASHBOARD_TEMPLATE = f"""
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Admin Dashboard - MNM Software</title>
    {COMMON_STYLES}
</head>
<body>
    <div class="animated-bg"></div>

    <div id="loading-overlay">
        <div class="spinner-container">
            <div class="spinner" id="spinner-anim"></div>
            <div class="spinner-inner"></div>
        </div>
        <div class="checkmark-container" id="success-anim">
            <div class="checkmark-circle"></div>
            <div class="anim-text">Successful!</div>
        </div>
        <div class="fail-container" id="fail-anim">
            <div class="fail-circle"></div>
            <div class="anim-text">Action Failed!</div>
        </div>
        <div class="loading-text" id="loading-text">Processing Request...</div>
    </div>

    <div id="flash-container">
        {{% with messages = get_flashed_messages(with_categories=true) %}}
            {{% if messages %}}
                {{% for category, message in messages %}}
                    <div class="flash-message flash-{{{{ category if category != 'message' else 'info' }}}}" role="alert">
                        {{% if category == 'success' %}}
                            <i class="fas fa-check-circle"></i>
                        {{% elif category == 'error' %}}
                            <i class="fas fa-times-circle"></i>
                        {{% elif category == 'warning' %}}
                            <i class="fas fa-exclamation-triangle"></i>
                        {{% else %}}
                            <i class="fas fa-info-circle"></i>
                        {{% endif %}}
                        <span>{{{{ message }}}}</span>
                        <button class="flash-close" onclick="this.parentElement.remove()">
                            <i class="fas fa-times"></i>
                        </button>
                    </div>
                {{% endfor %}}
            {{% endif %}}
        {{% endwith %}}
    </div>

    <div class="mobile-toggle" onclick="document.querySelector('.sidebar').classList.toggle('active')">
        <i class="fas fa-bars"></i>
    </div>

    <div class="sidebar">
        <div class="brand-logo">
            <i class="fas fa-layer-group"></i> 
            MNM<span>Software</span>
        </div>
        <div class="nav-menu">
            <div class="nav-link active" onclick="showSection('dashboard', this)">
                <i class="fas fa-home"></i> Dashboard
                <span class="nav-badge">Live</span>
            </div>
            <div class="nav-link" onclick="showSection('analytics', this)">
                <i class="fas fa-chart-pie"></i> Closing Report
            </div>
            <a href="/admin/accessories" class="nav-link">
                <i class="fas fa-database"></i> Accessories Challan
            </a>
            <div class="nav-link" onclick="showSection('help', this)">
                <i class="fas fa-file-invoice"></i> PO Generator
            </div>
            <div class="nav-link" onclick="showSection('settings', this)">
                <i class="fas fa-users-cog"></i> User Manage
            </div>
            <a href="/logout" class="nav-link" style="color: var(--accent-red); margin-top: 20px;">
                <i class="fas fa-sign-out-alt"></i> Sign Out
            </a>
        </div>
        <div class="sidebar-footer">
            <i class="fas fa-code" style="margin-right: 5px;"></i> Powered by Mehedi Hasan
        </div>
    </div>

    <div class="main-content">
        
        <!-- DASHBOARD SECTION -->
        <div id="section-dashboard">
            <div class="header-section">
                <div>
                    <div class="page-title">Main Dashboard</div>
                    <div class="page-subtitle">Lifetime Overview & Real-time Analytics</div>
                </div>
                <div class="status-badge">
                    <div class="status-dot"></div>
                    <span>System Online</span>
                </div>
            </div>

            <div class="stats-grid">
                <div class="card stat-card">
                    <div class="stat-icon"><i class="fas fa-file-export"></i></div>
                    <div class="stat-info">
                        <h3 class="count-up" data-target="{{{{ stats.closing.count }}}}">0</h3>
                        <p>Lifetime Closing</p>
                    </div>
                </div>
                <div class="card stat-card">
                    <div class="stat-icon" style="background: linear-gradient(145deg, rgba(139, 92, 246, 0.15), rgba(139, 92, 246, 0.05));">
                        <i class="fas fa-boxes" style="color: var(--accent-purple);"></i>
                    </div>
                    <div class="stat-info">
                        <h3 class="count-up" data-target="{{{{ stats.accessories.count }}}}">0</h3>
                        <p>Lifetime Accessories</p>
                    </div>
                </div>
                <div class="card stat-card">
                    <div class="stat-icon" style="background: linear-gradient(145deg, rgba(16, 185, 129, 0.15), rgba(16, 185, 129, 0.05));">
                        <i class="fas fa-file-pdf" style="color: var(--accent-green);"></i>
                    </div>
                    <div class="stat-info">
                        <h3 class="count-up" data-target="{{{{ stats.po.count }}}}">0</h3>
                        <p>Lifetime PO Sheets</p>
                    </div>
                </div>
                <div class="card stat-card">
                    <div class="stat-icon" style="background: linear-gradient(145deg, rgba(59, 130, 246, 0.15), rgba(59, 130, 246, 0.05));">
                        <i class="fas fa-users" style="color: var(--accent-blue);"></i>
                    </div>
                    <div class="stat-info">
                        <h3 class="count-up" data-target="{{{{ stats.users.count }}}}">0</h3>
                        <p>Total Users</p>
                    </div>
                </div>
            </div>

            <div class="dashboard-grid-2">
                <div class="card">
                    <div class="section-header">
                        <span>Daily Activity Chart</span>
                        <div class="realtime-indicator">
                            <div class="realtime-dot"></div>
                            <span>Real-time</span>
                        </div>
                    </div>
                    <div class="chart-container" style="height: 320px;">
                        <canvas id="mainChart"></canvas>
                    </div>
                </div>
                <div class="card">
                    <div class="section-header">
                        <span>24h Summary</span>
                        <i class="fas fa-clock" style="color: var(--accent-orange);"></i>
                    </div>
                    
                    <div class="progress-item">
                        <div class="progress-header">
                            <span>Closing Report</span>
                            <span class="progress-value">{{{{ stats.counts_24h.closing }}}} today</span>
                        </div>
                        <div class="progress-bar-container">
                            <div class="progress-bar-fill progress-orange" style="width: {{{{ (stats.counts_24h.closing / (stats.counts_24h.total + 1)) * 100 }}}}%;"></div>
                        </div>
                    </div>
                    
                    <div class="progress-item">
                        <div class="progress-header">
                            <span>Accessories</span>
                            <span class="progress-value">{{{{ stats.counts_24h.acc }}}} challans</span>
                        </div>
                        <div class="progress-bar-container">
                            <div class="progress-bar-fill progress-purple" style="width: {{{{ (stats.counts_24h.acc / (stats.counts_24h.total + 1)) * 100 }}}}%;"></div>
                        </div>
                    </div>
                    
                    <div class="progress-item">
                        <div class="progress-header">
                            <span>PO Generator</span>
                            <span class="progress-value">{{{{ stats.counts_24h.po }}}} files</span>
                        </div>
                        <div class="progress-bar-container">
                            <div class="progress-bar-fill progress-green" style="width: {{{{ (stats.counts_24h.po / (stats.counts_24h.total + 1)) * 100 }}}}%;"></div>
                        </div>
                    </div>

                    <div style="margin-top: 25px; padding: 15px; background: rgba(255,122,0,0.1); border-radius: 12px; text-align: center;">
                        <div style="font-size: 32px; font-weight: 800; color: var(--accent-orange);">{{{{ stats.counts_24h.total }}}}</div>
                        <div style="font-size: 11px; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 1px;">Total Actions (24h)</div>
                    </div>
                </div>
            </div>

            <!-- Activity Log with Search -->
            <div class="card">
                <div class="section-header">
                    <span>Activity Log (Last 24 Hours)</span>
                    <span class="table-badge" style="background: var(--accent-orange); color: white;">{{{{ stats.counts_24h.total }}}} Actions</span>
                </div>

                <!-- Search Box -->
                <div class="search-box">
                    <input type="text" id="historySearch" placeholder="Search by ref, user, type..." onkeyup="filterHistory()">
                    <i class="fas fa-search"></i>
                </div>

                <!-- Filter Tabs -->
                <div class="history-tabs">
                    <div class="history-tab active" onclick="filterByType('all', this)">
                        All <span class="tab-count">{{{{ stats.counts_24h.total }}}}</span>
                    </div>
                    <div class="history-tab" onclick="filterByType('Closing Report', this)">
                        <i class="fas fa-file-export" style="margin-right: 5px;"></i> Closing <span class="tab-count">{{{{ stats.counts_24h.closing }}}}</span>
                    </div>
                    <div class="history-tab" onclick="filterByType('Accessories', this)">
                        <i class="fas fa-boxes" style="margin-right: 5px;"></i> Accessories <span class="tab-count">{{{{ stats.counts_24h.acc }}}}</span>
                    </div>
                    <div class="history-tab" onclick="filterByType('PO Sheet', this)">
                        <i class="fas fa-file-pdf" style="margin-right: 5px;"></i> PO Sheet <span class="tab-count">{{{{ stats.counts_24h.po }}}}</span>
                    </div>
                </div>

                <div style="overflow-x: auto; max-height: 400px; overflow-y: auto;">
                    <table class="dark-table" id="historyTable">
                        <thead style="position: sticky; top: 0; background: var(--bg-card); z-index: 10;">
                            <tr>
                                <th>Time</th>
                                <th>User</th>
                                <th>Action</th>
                                <th>Reference</th>
                                <th>Details</th>
                            </tr>
                        </thead>
                        <tbody id="historyBody">
                            {{% for log in stats.history_24h %}}
                            <tr class="history-row" data-type="{{{{ log.type }}}}" style="animation: fadeInUp 0.3s ease-out {{{{ loop.index * 0.03 }}}}s backwards;">
                                <td>
                                    <div class="time-badge">
                                        <i class="far fa-clock"></i>
                                        {{{{ log.time }}}}
                                    </div>
                                </td>
                                <td style="font-weight: 600; color: white;">{{{{ log.user }}}}</td>
                                <td>
                                    <span class="table-badge" style="
                                        {{% if log.type == 'Closing Report' %}}
                                        background: rgba(255, 122, 0, 0.15); color: var(--accent-orange);
                                        {{% elif log.type == 'PO Sheet' %}}
                                        background: rgba(16, 185, 129, 0.15); color: var(--accent-green);
                                        {{% elif log.type == 'Accessories' %}}
                                        background: rgba(139, 92, 246, 0.15); color: var(--accent-purple);
                                        {{% else %}}
                                        background: rgba(59, 130, 246, 0.15); color: var(--accent-blue);
                                        {{% endif %}}
                                    ">
                                        {{% if log.type == 'Closing Report' %}}
                                            <i class="fas fa-file-export" style="margin-right: 5px;"></i>
                                        {{% elif log.type == 'PO Sheet' %}}
                                            <i class="fas fa-file-pdf" style="margin-right: 5px;"></i>
                                        {{% elif log.type == 'Accessories' %}}
                                            <i class="fas fa-boxes" style="margin-right: 5px;"></i>
                                        {{% endif %}}
                                        {{{{ log.type }}}}
                                    </span>
                                </td>
                                <td style="color: var(--accent-orange); font-weight: 600;">{{{{ log.ref if log.ref else '-' }}}}</td>
                                <td style="color: var(--text-secondary); font-size: 12px;">{{{{ log.info if log.info else log.date }}}}</td>
                            </tr>
                            {{% else %}}
                            <tr id="emptyRow">
                                <td colspan="5" style="text-align: center; padding: 50px; color: var(--text-secondary);">
                                    <i class="fas fa-inbox" style="font-size: 50px; opacity: 0.2; margin-bottom: 15px; display: block;"></i>
                                    <h4 style="color: white; margin-bottom: 8px;">No Activity Yet</h4>
                                    No actions recorded in the last 24 hours.
                                </td>
                            </tr>
                            {{% endfor %}}
                        </tbody>
                    </table>
                    
                    <div class="no-results" id="noResults" style="display: none;">
                        <i class="fas fa-search"></i>
                        <h4>No Results Found</h4>
                        <p>Try different keywords or clear the search</p>
                    </div>
                </div>
            </div>
        </div>

        <!-- CLOSING REPORT SECTION -->
        <div id="section-analytics" style="display: none;">
            <div class="header-section">
                <div>
                    <div class="page-title">Closing Report</div>
                    <div class="page-subtitle">Generate production closing reports</div>
                </div>
            </div>
            <div class="card" style="max-width: 550px; margin: 0 auto; margin-top: 30px;">
                <div class="section-header">
                    <span><i class="fas fa-magic" style="margin-right: 10px; color: var(--accent-orange);"></i>Generate Report</span>
                </div>
                <form action="/generate-report" method="post" onsubmit="return showLoading()">
                    <div class="input-group">
                        <label><i class="fas fa-bookmark" style="margin-right: 5px;"></i> INTERNAL REF NO</label>
                        <input type="text" name="ref_no" placeholder="e.g. IB-12345 or Booking-123" required>
                    </div>
                    <button type="submit">
                        <i class="fas fa-bolt" style="margin-right: 10px;"></i> Generate Report
                    </button>
                </form>
            </div>
        </div>

        <!-- PO GENERATOR SECTION -->
        <div id="section-help" style="display: none;">
            <div class="header-section">
                <div>
                    <div class="page-title">PO Sheet Generator</div>
                    <div class="page-subtitle">Process and generate PO summary sheets</div>
                </div>
            </div>
            <div class="card" style="max-width: 650px; margin: 0 auto; margin-top: 30px;">
                <div class="section-header">
                    <span><i class="fas fa-file-pdf" style="margin-right: 10px; color: var(--accent-green);"></i>Upload PDF Files</span>
                </div>
                <form action="/generate-po-report" method="post" enctype="multipart/form-data" onsubmit="return showLoading()">
                    <div class="upload-zone" id="uploadZone" onclick="document.getElementById('file-upload').click()">
                        <input type="file" name="pdf_files" multiple accept=".pdf" required style="display: none;" id="file-upload">
                        <div class="upload-icon">
                            <i class="fas fa-cloud-upload-alt"></i>
                        </div>
                        <div class="upload-text">Click or Drag to Upload PDF Files</div>
                        <div class="upload-hint">Supports multiple PDF files (Booking + PO files)</div>
                        <div id="file-count">No files selected</div>
                    </div>
                    <button type="submit" style="margin-top: 25px; background: linear-gradient(135deg, #10B981 0%, #34D399 100%);">
                        <i class="fas fa-cogs" style="margin-right: 10px;"></i> Process Files
                    </button>
                </form>
            </div>
        </div>

        <!-- USER MANAGEMENT SECTION -->
        <div id="section-settings" style="display:none;">
            <div class="header-section">
                <div>
                    <div class="page-title">User Management</div>
                    <div class="page-subtitle">Manage user accounts and permissions</div>
                </div>
            </div>
            <div class="dashboard-grid-2">
                <div class="card">
                    <div class="section-header">
                        <span>User Directory</span>
                        <span class="table-badge" style="background: var(--accent-orange); color: white;">{{{{ stats.users.count }}}} Users</span>
                    </div>
                    <div id="userTableContainer" style="max-height: 450px; overflow-y: auto;">
                        <div style="text-align: center; padding: 40px; color: var(--text-secondary);">
                            <i class="fas fa-spinner fa-spin" style="font-size: 30px; margin-bottom: 10px;"></i>
                            <div>Loading users...</div>
                        </div>
                    </div>
                </div>
                <div class="card">
                    <div class="section-header">
                        <span>Create / Edit User</span>
                        <i class="fas fa-user-plus" style="color: var(--accent-orange);"></i>
                    </div>
                    <form id="userForm">
                        <input type="hidden" id="action_type" value="create">
                        <div class="input-group">
                            <label><i class="fas fa-user" style="margin-right: 5px;"></i> USERNAME</label>
                            <input type="text" id="new_username" required placeholder="Enter username">
                        </div>
                        <div class="input-group">
                            <label><i class="fas fa-key" style="margin-right: 5px;"></i> PASSWORD</label>
                            <input type="text" id="new_password" required placeholder="Enter password">
                        </div>
                        <div class="input-group">
                            <label><i class="fas fa-shield-alt" style="margin-right: 5px;"></i> PERMISSIONS</label>
                            <div style="display: flex; gap: 10px; flex-wrap: wrap; margin-top: 5px;">
                                <label class="perm-checkbox">
                                    <input type="checkbox" id="perm_closing" checked>
                                    <span>Closing</span>
                                </label>
                                <label class="perm-checkbox">
                                    <input type="checkbox" id="perm_po">
                                    <span>PO Sheet</span>
                                </label>
                                <label class="perm-checkbox">
                                    <input type="checkbox" id="perm_acc">
                                    <span>Accessories</span>
                                </label>
                            </div>
                        </div>
                        <button type="button" onclick="handleUserSubmit()" id="saveUserBtn">
                            <i class="fas fa-save" style="margin-right: 10px;"></i> Save User
                        </button>
                        <button type="button" onclick="resetForm()" style="margin-top: 12px; background: rgba(255,255,255,0.05); border: 1px solid var(--border-color);">
                            <i class="fas fa-undo" style="margin-right: 10px;"></i> Reset Form
                        </button>
                    </form>
                </div>
            </div>
        </div>
    </div>

    <script>
        let currentFilter = 'all';
        let allHistoryRows = [];

        document.addEventListener('DOMContentLoaded', function() {{
            allHistoryRows = Array.from(document.querySelectorAll('.history-row'));
        }});

        function showSection(id, element) {{
            ['dashboard', 'analytics', 'help', 'settings'].forEach(sid => {{
                document.getElementById('section-' + sid).style.display = 'none';
            }});
            document.getElementById('section-' + id).style.display = 'block';
            
            if (element) {{
                document.querySelectorAll('.nav-link').forEach(el => el.classList.remove('active'));
                element.classList.add('active');
            }}
            
            if (id === 'settings') loadUsers();
            if (window.innerWidth < 1024) document.querySelector('.sidebar').classList.remove('active');
        }}
        
        function filterHistory() {{
            const searchTerm = document.getElementById('historySearch').value.toLowerCase();
            let visibleCount = 0;

            allHistoryRows.forEach(row => {{
                const text = row.textContent.toLowerCase();
                const type = row.getAttribute('data-type');
                
                const matchesSearch = text.includes(searchTerm);
                const matchesFilter = currentFilter === 'all' || type === currentFilter;
                
                if (matchesSearch && matchesFilter) {{
                    row.style.display = '';
                    visibleCount++;
                }} else {{
                    row.style.display = 'none';
                }}
            }});

            const noResults = document.getElementById('noResults');
            const emptyRow = document.getElementById('emptyRow');
            
            if (visibleCount === 0 && allHistoryRows.length > 0) {{
                noResults.style.display = 'block';
                if (emptyRow) emptyRow.style.display = 'none';
            }} else {{
                noResults.style.display = 'none';
            }}
        }}

        function filterByType(type, element) {{
            currentFilter = type;
            document.querySelectorAll('.history-tab').forEach(tab => tab.classList.remove('active'));
            element.classList.add('active');
            filterHistory();
        }}

        const fileUpload = document.getElementById('file-upload');
        const uploadZone = document.getElementById('uploadZone');
        
        if (fileUpload) {{
            fileUpload.addEventListener('change', function() {{
                const count = this.files.length;
                document.getElementById('file-count').innerHTML = count > 0 
                    ? `<i class="fas fa-check-circle" style="margin-right: 5px; color: var(--accent-green);"></i>${{count}} file(s) selected`
                    : 'No files selected';
            }});
            uploadZone.addEventListener('dragover', (e) => {{
                e.preventDefault();
                uploadZone.classList.add('dragover');
            }});
            uploadZone.addEventListener('dragleave', () => {{
                uploadZone.classList.remove('dragover');
            }});
            uploadZone.addEventListener('drop', (e) => {{
                e.preventDefault();
                uploadZone.classList.remove('dragover');
                fileUpload.files = e.dataTransfer.files;
                fileUpload.dispatchEvent(new Event('change'));
            }});
        }}
        
        const ctx = document.getElementById('mainChart').getContext('2d');
        const gradientOrange = ctx.createLinearGradient(0, 0, 0, 300);
        gradientOrange.addColorStop(0, 'rgba(255, 122, 0, 0.5)');
        gradientOrange.addColorStop(1, 'rgba(255, 122, 0, 0.0)');
        const gradientPurple = ctx.createLinearGradient(0, 0, 0, 300);
        gradientPurple.addColorStop(0, 'rgba(139, 92, 246, 0.5)');
        gradientPurple.addColorStop(1, 'rgba(139, 92, 246, 0.0)');
        const gradientGreen = ctx.createLinearGradient(0, 0, 0, 300);
        gradientGreen.addColorStop(0, 'rgba(16, 185, 129, 0.5)');
        gradientGreen.addColorStop(1, 'rgba(16, 185, 129, 0.0)');
        
        new Chart(ctx, {{
            type: 'line',
            data: {{
                labels: {{{{ stats.chart.labels | tojson }}}},
                datasets: [
                    {{
                        label: 'Closing',
                        data: {{{{ stats.chart.closing | tojson }}}},
                        borderColor: '#FF7A00',
                        backgroundColor: gradientOrange,
                        tension: 0.4,
                        fill: true,
                        pointRadius: 4,
                        borderWidth: 3
                    }},
                    {{
                        label: 'Accessories',
                        data: {{{{ stats.chart.acc | tojson }}}},
                        borderColor: '#8B5CF6',
                        backgroundColor: gradientPurple,
                        tension: 0.4,
                        fill: true,
                        pointRadius: 4,
                        borderWidth: 3
                    }},
                    {{
                        label: 'PO Sheets',
                        data: {{{{ stats.chart.po | tojson }}}},
                        borderColor: '#10B981',
                        backgroundColor: gradientGreen,
                        tension: 0.4,
                        fill: true,
                        pointRadius: 4,
                        borderWidth: 3
                    }}
                ]
            }},
            options: {{
                plugins: {{
                    legend: {{
                        display: true,
                        position: 'top',
                        labels: {{ color: '#8b8b9e', font: {{ size: 11 }}, usePointStyle: true, padding: 15 }}
                    }}
                }},
                scales: {{
                    x: {{ grid: {{ display: false }}, ticks: {{ color: '#8b8b9e', font: {{ size: 10 }} }} }},
                    y: {{ grid: {{ color: 'rgba(255,255,255,0.03)' }}, ticks: {{ color: '#8b8b9e' }}, beginAtZero: true }}
                }},
                responsive: true,
                maintainAspectRatio: false
            }}
        }});

        function animateCountUp() {{
            document.querySelectorAll('.count-up').forEach(counter => {{
                const target = parseInt(counter.getAttribute('data-target'));
                const duration = 2000;
                let start = 0;
                const increment = target / (duration / 16);
                
                const updateCounter = () => {{
                    start += increment;
                    if (start < target) {{
                        counter.textContent = Math.floor(start).toLocaleString();
                        requestAnimationFrame(updateCounter);
                    }} else {{
                        counter.textContent = target.toLocaleString();
                    }}
                }};
                updateCounter();
            }});
        }}
        setTimeout(animateCountUp, 500);

        function showLoading() {{
            const overlay = document.getElementById('loading-overlay');
            overlay.style.display = 'flex';
            document.getElementById('spinner-anim').parentElement.style.display = 'block';
            document.getElementById('success-anim').style.display = 'none';
            document.getElementById('fail-anim').style.display = 'none';
            document.getElementById('loading-text').style.display = 'block';
            return true;
        }}

        function showSuccess() {{
            document.getElementById('spinner-anim').parentElement.style.display = 'none';
            document.getElementById('success-anim').style.display = 'block';
            document.getElementById('loading-text').style.display = 'none';
            setTimeout(() => {{ document.getElementById('loading-overlay').style.display = 'none'; }}, 1500);
        }}

        function loadUsers() {{
            fetch('/admin/get-users')
                .then(res => res.json())
                .then(data => {{
                    let html = '<table class="dark-table"><thead><tr><th>User</th><th>Last Seen</th><th style="text-align:right;">Actions</th></tr></thead><tbody>';
                    for (const [u, d] of Object.entries(data)) {{
                        const roleStyle = d.role === 'admin' ? 'background: rgba(255, 122, 0, 0.1); color: var(--accent-orange);' : 'background: rgba(139, 92, 246, 0.1); color: var(--accent-purple);';
                        html += `<tr>
                            <td>
                                <div style="font-weight: 600;">${{u}}</div>
                                <span class="table-badge" style="${{roleStyle}}">${{d.role}}</span>
                            </td>
                            <td><span class="table-badge">${{d.last_login}}</span></td>
                            <td style="text-align:right;">
                                ${{d.role !== 'admin' ? `
                                    <div class="action-cell">
                                        <button class="action-btn btn-edit" onclick="editUser('${{u}}', '${{d.password}}', '${{d.permissions.join(',')}}')">
                                            <i class="fas fa-edit"></i>
                                        </button> 
                                        <button class="action-btn btn-del" onclick="deleteUser('${{u}}')">
                                            <i class="fas fa-trash"></i>
                                        </button>
                                    </div>
                                ` : '<i class="fas fa-shield-alt" style="color: var(--accent-orange); opacity: 0.5;"></i>'}}
                            </td>
                        </tr>`;
                    }}
                    document.getElementById('userTableContainer').innerHTML = html + '</tbody></table>';
                }});
        }}
        
        function handleUserSubmit() {{
            const u = document.getElementById('new_username').value;
            const p = document.getElementById('new_password').value;
            const a = document.getElementById('action_type').value;
            let perms = [];
            if (document.getElementById('perm_closing').checked) perms.push('closing');
            if (document.getElementById('perm_po').checked) perms.push('po_sheet');
            if (document.getElementById('perm_acc').checked) perms.push('accessories');
            
            showLoading();
            fetch('/admin/save-user', {{
                method: 'POST',
                headers: {{'Content-Type': 'application/json'}},
                body: JSON.stringify({{ username: u, password: p, permissions: perms, action_type: a }})
            }})
            .then(r => r.json())
            .then(d => {{
                if (d.status === 'success') {{
                    showSuccess();
                    loadUsers();
                    resetForm();
                }} else {{
                    alert(d.message);
                    document.getElementById('loading-overlay').style.display = 'none';
                }}
            }});
        }}
        
        function editUser(u, p, permsStr) {{
            document.getElementById('new_username').value = u;
            document.getElementById('new_username').readOnly = true;
            document.getElementById('new_password').value = p;
            document.getElementById('action_type').value = 'update';
            document.getElementById('saveUserBtn').innerHTML = '<i class="fas fa-sync" style="margin-right: 10px;"></i> Update User';
            const pArr = permsStr.split(',');
            document.getElementById('perm_closing').checked = pArr.includes('closing');
            document.getElementById('perm_po').checked = pArr.includes('po_sheet');
            document.getElementById('perm_acc').checked = pArr.includes('accessories');
        }}
        
        function resetForm() {{
            document.getElementById('userForm').reset();
            document.getElementById('action_type').value = 'create';
            document.getElementById('saveUserBtn').innerHTML = '<i class="fas fa-save" style="margin-right: 10px;"></i> Save User';
            document.getElementById('new_username').readOnly = false;
            document.getElementById('perm_closing').checked = true;
        }}
        
        function deleteUser(u) {{
            if (confirm('Are you sure you want to delete "' + u + '"?')) {{
                fetch('/admin/delete-user', {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{ username: u }})
                }}).then(() => loadUsers());
            }}
        }}

        setTimeout(function() {{
            document.querySelectorAll('.flash-message').forEach(function(msg) {{
                msg.style.animation = 'flashSlideOut 0.5s ease-out forwards';
                setTimeout(function() {{ msg.remove(); }}, 500);
            }});
        }}, 5000);
    </script>
</body>
</html>
"""
# ==============================================================================
# USER DASHBOARD TEMPLATE
# ==============================================================================

USER_DASHBOARD_TEMPLATE = f"""
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Dashboard - MNM Software</title>
    {COMMON_STYLES}
</head>
<body>
    <div class="animated-bg"></div>
    
    <div id="loading-overlay">
        <div class="spinner-container">
            <div class="spinner"></div>
            <div class="spinner-inner"></div>
        </div>
        <div class="loading-text">Processing...</div>
    </div>
    
    <div id="flash-container">
        {{% with messages = get_flashed_messages(with_categories=true) %}}
            {{% if messages %}}
                {{% for category, message in messages %}}
                    <div class="flash-message flash-{{{{ category if category != 'message' else 'info' }}}}" role="alert">
                        {{% if category == 'success' %}}
                            <i class="fas fa-check-circle"></i>
                        {{% elif category == 'error' %}}
                            <i class="fas fa-times-circle"></i>
                        {{% elif category == 'warning' %}}
                            <i class="fas fa-exclamation-triangle"></i>
                        {{% else %}}
                            <i class="fas fa-info-circle"></i>
                        {{% endif %}}
                        <span>{{{{ message }}}}</span>
                        <button class="flash-close" onclick="this.parentElement.remove()">
                            <i class="fas fa-times"></i>
                        </button>
                    </div>
                {{% endfor %}}
            {{% endif %}}
        {{% endwith %}}
    </div>
    
    <div class="sidebar">
        <div class="brand-logo">
            <i class="fas fa-layer-group"></i> 
            MNM<span>Software</span>
        </div>
        <div class="nav-menu">
            <div class="nav-link active">
                <i class="fas fa-home"></i> Home
            </div>
            <a href="/logout" class="nav-link" style="color: var(--accent-red); margin-top: auto;">
                <i class="fas fa-sign-out-alt"></i> Sign Out
            </a>
        </div>
        <div class="sidebar-footer">
            <i class="fas fa-code" style="margin-right: 5px;"></i> Powered by Mehedi Hasan
        </div>
    </div>
    
    <div class="main-content">
        <div class="header-section">
            <div>
                <div class="page-title">Welcome, {{{{ session.user }}}}!</div>
                <div class="page-subtitle">Your assigned production modules</div>
            </div>
            <div class="status-badge">
                <div class="status-dot"></div>
                <span>Online</span>
            </div>
        </div>

        <div class="stats-grid">
            {{% if 'closing' in session.permissions %}}
            <div class="card" style="animation: fadeInUp 0.5s ease-out 0.1s backwards;">
                <div class="section-header">
                    <span><i class="fas fa-file-export" style="margin-right: 10px; color: var(--accent-orange);"></i>Closing Report</span>
                </div>
                <p style="color: var(--text-secondary); margin-bottom: 20px; font-size: 14px; line-height: 1.6;">
                    Generate production closing reports with real-time data. 
                </p>
                <form action="/generate-report" method="post" onsubmit="showLoading()">
                    <div class="input-group">
                        <label>BOOKING REF NO</label>
                        <input type="text" name="ref_no" required placeholder="Enter Booking Reference">
                    </div>
                    <button type="submit">
                        <i class="fas fa-magic" style="margin-right: 8px;"></i> Generate
                    </button>
                </form>
            </div>
            {{% endif %}}
            
            {{% if 'po_sheet' in session.permissions %}}
            <div class="card" style="animation: fadeInUp 0.5s ease-out 0.2s backwards;">
                <div class="section-header">
                    <span><i class="fas fa-file-pdf" style="margin-right: 10px; color: var(--accent-green);"></i>PO Sheet</span>
                </div>
                <p style="color: var(--text-secondary); margin-bottom: 20px; font-size: 14px; line-height: 1.6;">
                    Process PDF files and generate PO summary reports.
                </p>
                <form action="/generate-po-report" method="post" enctype="multipart/form-data" onsubmit="showLoading()">
                    <div class="input-group">
                        <label>PDF FILES</label>
                        <input type="file" name="pdf_files" multiple accept=".pdf" required style="padding: 12px;">
                    </div>
                    <button type="submit" style="background: linear-gradient(135deg, #10B981 0%, #34D399 100%);">
                        <i class="fas fa-cogs" style="margin-right: 8px;"></i> Process Files
                    </button>
                </form>
            </div>
            {{% endif %}}
            
            {{% if 'accessories' in session.permissions %}}
            <div class="card" style="animation: fadeInUp 0.5s ease-out 0.3s backwards;">
                <div class="section-header">
                    <span><i class="fas fa-boxes" style="margin-right: 10px; color: var(--accent-purple);"></i>Accessories</span>
                </div>
                <p style="color: var(--text-secondary); margin-bottom: 20px; font-size: 14px; line-height: 1.6;">
                    Manage challans, entries and delivery history for accessories.
                </p>
                <a href="/admin/accessories">
                    <button style="background: linear-gradient(135deg, #8B5CF6 0%, #A78BFA 100%);">
                        <i class="fas fa-external-link-alt" style="margin-right: 8px;"></i> Open Dashboard
                    </button>
                </a>
            </div>
            {{% endif %}}
        </div>
    </div>
    
    <script>
        function showLoading() {{
            document.getElementById('loading-overlay').style.display = 'flex';
            return true;
        }}
        
        setTimeout(function() {{
            document.querySelectorAll('.flash-message').forEach(function(msg) {{
                msg.style.animation = 'flashSlideOut 0.5s ease-out forwards';
                setTimeout(function() {{ msg.remove(); }}, 500);
            }});
        }}, 5000);

        const style = document.createElement('style');
        style.textContent = `
            @keyframes flashSlideOut {{
                from {{ opacity: 1; transform: translateY(0); }}
                to {{ opacity: 0; transform: translateY(-30px); }}
            }}
        `;
        document.head.appendChild(style);
    </script>
</body>
</html>
"""

# ==============================================================================
# NEW PO REPORT TEMPLATE (আপনার দেওয়া কোড থেকে হুবহু - Light Theme)
# ==============================================================================

PO_REPORT_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PO Report - Cotton Clothing BD</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root {
            --primary: #1e40af;
            --primary-light: #3b82f6;
            --dark: #0f172a;
            --dark-light: #1e293b;
            --gray-50: #f8fafc;
            --gray-100: #f1f5f9;
            --gray-200: #e2e8f0;
            --gray-300: #cbd5e1;
            --gray-600: #475569;
            --gray-800: #1e293b;
            --success: #059669;
            --success-light: #d1fae5;
            --warning: #d97706;
            --warning-light: #fef3c7;
        }
        
        * { font-family: 'Inter', sans-serif; }
        
        body { 
            background: var(--gray-100);
            min-height: 100vh;
            padding: 30px 0; 
        }
        
        .container { max-width: 1200px; }
        
        .company-header { 
            background: white;
            border-radius: 12px;
            padding: 24px 32px;
            margin-bottom: 20px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            border-left: 4px solid var(--primary);
        }
        
        .header-content {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .company-name { 
            font-size: 1.5rem; 
            font-weight: 800; 
            color: var(--dark);
            letter-spacing: -0.5px;
            margin-bottom: 4px;
        }
        
        .report-title { 
            font-size: 0.85rem; 
            color: var(--gray-600);
            font-weight: 500;
            text-transform: uppercase;
            letter-spacing: 1px;
        }
        
        .date-box { text-align: right; }
        
        .date-label {
            font-size: 0.7rem;
            color: var(--gray-600);
            text-transform: uppercase;
            letter-spacing: 1px;
            font-weight: 600;
        }
        
        .date-value {
            font-size: 1.1rem;
            font-weight: 700;
            color: var(--dark);
        }
        
        .info-section {
            display: grid;
            grid-template-columns: 1fr 220px;
            gap: 20px;
            margin-bottom: 20px;
        }
        
        .info-grid { 
            background: white;
            border-radius: 12px;
            padding: 24px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 20px;
        }
        
        .info-item {
            padding: 12px 16px;
            background: var(--gray-50);
            border-radius: 8px;
            border-left: 3px solid var(--primary-light);
        }
        
        .info-item.booking-highlight {
            background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
            border-left: 4px solid #f59e0b;
            box-shadow: 0 2px 8px rgba(245, 158, 11, 0.25);
        }
        
        .info-item.booking-highlight .info-value {
            color: #92400e;
            font-size: 1.05rem;
        }
        
        .info-label { 
            font-size: 0.65rem;
            font-weight: 700;
            color: var(--gray-600);
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-bottom: 4px;
        }
        
        .info-value { 
            font-size: 0.95rem;
            font-weight: 700;
            color: var(--dark);
        }

        .grand-total-box { 
            background: linear-gradient(135deg, var(--dark) 0%, var(--dark-light) 100%);
            color: white; 
            padding: 24px;
            border-radius: 12px;
            text-align: center;
            display: flex;
            flex-direction: column;
            justify-content: center;
            box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        }
        
        .grand-total-label { 
            font-size: 0.7rem;
            text-transform: uppercase;
            letter-spacing: 2px;
            font-weight: 600;
            opacity: 0.8;
            margin-bottom: 8px;
        }
        
        .grand-total-value { 
            font-size: 2.2rem;
            font-weight: 800;
            line-height: 1;
        }
        
        .grand-total-unit {
            font-size: 0.75rem;
            opacity: 0.7;
            margin-top: 6px;
            font-weight: 500;
        }

        .table-card { 
            background: white;
            border-radius: 12px;
            margin-bottom: 20px;
            overflow: hidden;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }
        
        .color-header { 
            background: var(--dark);
            color: white;
            padding: 14px 20px;
            font-size: 0.9rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 1px;
        }

        .table { margin-bottom: 0; width: 100%; }
        
        .table th { 
            background: var(--gray-100);
            color: var(--gray-800);
            font-weight: 700;
            font-size: 0.75rem;
            text-align: center;
            padding: 12px 10px;
            border-bottom: 2px solid var(--gray-200);
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        
        .table td { 
            text-align: center;
            vertical-align: middle;
            padding: 12px 10px;
            color: var(--dark);
            font-weight: 600;
            font-size: 0.9rem;
            border-bottom: 1px solid var(--gray-200);
        }
        
        .table tbody tr:hover td { background: var(--gray-50); }
        
        .order-col { 
            font-weight: 800 !important;
            background: var(--gray-50) !important;
            color: var(--primary) !important;
            border-right: 2px solid var(--gray-200) !important;
        }
        
        .total-col { 
            font-weight: 800 !important;
            background: var(--success-light) !important;
            color: var(--success) !important;
            border-left: 2px solid #a7f3d0 !important;
        }
        
        .total-col-header { 
            background: var(--success-light) !important;
            color: var(--success) !important;
            font-weight: 800 !important;
            border-left: 2px solid #a7f3d0 !important;
        }

        .table tbody tr.summary-row td { 
            background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%) !important;
            color: #92400e !important;
            font-weight: 800 !important;
            font-size: 1.0rem !important;
            border-top: 3px solid #f59e0b !important;
            border-bottom: 1px solid #fbbf24 !important;
            padding: 14px 10px !important;
        }
        
        .table tbody tr.summary-row:last-child td {
            background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%) !important;
            color: #1e40af !important;
            border-top: 3px solid #3b82f6 !important;
            border-bottom: none !important;
        }
        
        .summary-label { 
            text-align: right !important;
            padding-right: 20px !important;
            font-size: 0.85rem !important;
            text-transform: uppercase;
            letter-spacing: 1px;
        }

        .action-bar { 
            margin-bottom: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .btn-back {
            background: white;
            color: var(--dark);
            border: 2px solid var(--gray-200);
            border-radius: 8px;
            padding: 10px 24px;
            font-weight: 600;
            font-size: 0.9rem;
            transition: all 0.2s ease;
            text-decoration: none;
        }
        
        .btn-back:hover {
            border-color: var(--primary-light);
            color: var(--primary);
        }
        
        .btn-print { 
            background: var(--primary);
            color: white;
            border: none;
            border-radius: 8px;
            padding: 12px 28px;
            font-weight: 600;
            font-size: 0.9rem;
            transition: all 0.2s ease;
            box-shadow: 0 2px 8px rgba(30, 64, 175, 0.3);
            cursor: pointer;
        }
        
        .btn-print:hover {
            background: var(--primary-light);
            transform: translateY(-1px);
        }
        
        .footer-credit { 
            text-align: center;
            margin-top: 30px;
            padding: 16px;
            background: white;
            border-radius: 8px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            font-size: 0.85rem;
            color: var(--gray-600);
        }
        
        .footer-credit strong { color: var(--primary); }

        @media print {
            @page { margin: 10mm; size: portrait; }
            body { background: white !important; padding: 0 !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
            .container { max-width: 100% !important; padding: 0 !important; }
            .no-print { display: none !important; }
            .company-header { border-radius: 0 !important; box-shadow: none !important; border: 1px solid #000 !important; border-left: 4px solid #000 !important; margin-bottom: 10px !important; padding: 15px 20px !important; }
            .company-name { font-size: 1.3rem !important; }
            .info-section { margin-bottom: 10px !important; }
            .info-grid { border-radius: 0 !important; box-shadow: none !important; border: 1px solid #000 !important; padding: 15px !important; }
            .info-item { border-left: 3px solid #000 !important; }
            .info-item.booking-highlight { background: #fff8dc !important; border-left: 4px solid #000 !important; }
            .grand-total-box { border-radius: 0 !important; box-shadow: none !important; border: 2px solid #000 !important; background: #f0f0f0 !important; color: #000 !important; }
            .grand-total-box * { color: #000 !important; }
            .table-card { border-radius: 0 !important; box-shadow: none !important; border: 1px solid #000 !important; margin-bottom: 10px !important; break-inside: avoid; }
            .color-header { background: #e0e0e0 !important; color: #000 !important; padding: 10px 15px !important; font-size: 11pt !important; }
            .table th { background: #f5f5f5 !important; color: #000 !important; font-size: 9pt !important; padding: 8px 6px !important; border: 1px solid #000 !important; }
            .table td { font-size: 10pt !important; padding: 8px 6px !important; border: 1px solid #000 !important; }
            .order-col { background: #f8f8f8 !important; color: #000 !important; }
            .total-col, .total-col-header { background: #e8f5e9 !important; color: #000 !important; }
            .summary-row td { font-size: 10.5pt !important; border-top: 2px solid #000 !important; }
            .summary-row:first-of-type td { background: #fff8e1 !important; color: #000 !important; }
            .summary-row:last-of-type td { background: #e3f2fd !important; color: #000 !important; }
            .footer-credit { border-radius: 0 !important; box-shadow: none !important; border-top: 1px solid #000 !important; margin-top: 15px !important; padding: 10px !important; background: transparent !important; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="action-bar no-print">
            <a href="/" class="btn-back">← Back</a>
            <button onclick="window.print()" class="btn-print">Print Report</button>
        </div>

        <div class="company-header">
            <div class="header-content">
                <div>
                    <div class="company-name">Cotton Clothing BD Limited</div>
                    <div class="report-title">Purchase Order Summary Report</div>
                </div>
                <div class="date-box">
                    <div class="date-label">Report Date</div>
                    <div class="date-value" id="date"></div>
                </div>
            </div>
        </div>

        {% if message %}
            <div class="alert alert-warning text-center no-print">{{ message }}</div>
        {% endif %}

        {% if tables %}
            <div class="info-section">
                <div class="info-grid">
                    <div class="info-item">
                        <div class="info-label">Buyer</div>
                        <div class="info-value">{{ meta.buyer }}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">Season</div>
                        <div class="info-value">{{ meta.season }}</div>
                    </div>
                    <div class="info-item booking-highlight">
                        <div class="info-label">Booking No</div>
                        <div class="info-value">{{ meta.booking }}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">Department</div>
                        <div class="info-value">{{ meta.dept }}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">Style</div>
                        <div class="info-value">{{ meta.style }}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">Item</div>
                        <div class="info-value">{{ meta.item }}</div>
                    </div>
                </div>
                
                <div class="grand-total-box">
                    <div class="grand-total-label">Grand Total</div>
                    <div class="grand-total-value">{{ grand_total }}</div>
                    <div class="grand-total-unit">Pieces</div>
                </div>
            </div>

            {% for item in tables %}
                <div class="table-card">
                    <div class="color-header">Color: {{ item.color }}</div>
                    <div class="table-responsive">
                        {{ item.table | safe }}
                    </div>
                </div>
            {% endfor %}
            
            <div class="footer-credit">
                Report Generated by <strong>Mehedi Hasan</strong>
            </div>
        {% endif %}
    </div>

    <script>
        const d = new Date();
        document.getElementById('date').innerText = 
            String(d.getDate()).padStart(2,'0') + '-' + 
            String(d.getMonth()+1).padStart(2,'0') + '-' + 
            d.getFullYear();
    </script>
</body>
</html>
"""

# ==============================================================================
# ACCESSORIES SEARCH TEMPLATE
# ==============================================================================

ACCESSORIES_SEARCH_TEMPLATE = f"""
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Accessories Search - MNM Software</title>
    {COMMON_STYLES}
    <style>
        body {{
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }}
        
        .search-container {{
            position: relative;
            z-index: 10;
            width: 100%;
            max-width: 520px;
            padding: 20px;
        }}
        
        .search-card {{
            background: var(--gradient-card);
            border: 1px solid var(--border-color);
            border-radius: 24px;
            padding: 50px 45px;
            backdrop-filter: blur(20px);
            box-shadow: 0 25px 80px rgba(0, 0, 0, 0.5), 0 0 60px var(--accent-orange-glow);
            animation: cardAppear 0.6s cubic-bezier(0.4, 0, 0.2, 1);
        }}
        
        @keyframes cardAppear {{
            from {{ opacity: 0; transform: translateY(20px) scale(0.95); }}
            to {{ opacity: 1; transform: translateY(0) scale(1); }}
        }}
        
        .search-header {{
            text-align: center;
            margin-bottom: 40px;
        }}
        
        .search-icon {{
            width: 80px;
            height: 80px;
            background: linear-gradient(145deg, rgba(139, 92, 246, 0.2), rgba(139, 92, 246, 0.05));
            border-radius: 20px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            font-size: 36px;
            color: var(--accent-purple);
            margin-bottom: 20px;
            animation: iconFloat 3s ease-in-out infinite;
        }}
        
        @keyframes iconFloat {{
            0%, 100% {{ transform: translateY(0); }}
            50% {{ transform: translateY(-10px); }}
        }}
        
        .search-title {{
            font-size: 28px;
            font-weight: 800;
            color: white;
            margin-bottom: 8px;
        }}
        
        .search-subtitle {{
            color: var(--text-secondary);
            font-size: 14px;
        }}
        
        .nav-links {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid var(--border-color);
        }}
        
        .nav-links a {{
            color: var(--text-secondary);
            text-decoration: none;
            font-size: 13px;
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 6px;
            transition: var(--transition-smooth);
        }}
        
        .nav-links a:hover {{ color: var(--accent-orange); }}
        .nav-links a.logout {{ color: var(--accent-red); }}
        .nav-links a.logout:hover {{ color: #ff6b6b; }}
    </style>
</head>
<body>
    <div class="animated-bg"></div>
    
    <div id="flash-container">
        {{% with messages = get_flashed_messages(with_categories=true) %}}
            {{% if messages %}}
                {{% for category, message in messages %}}
                    <div class="flash-message flash-{{{{ category if category != 'message' else 'info' }}}}" role="alert">
                        {{% if category == 'success' %}}
                            <i class="fas fa-check-circle"></i>
                        {{% elif category == 'error' %}}
                            <i class="fas fa-times-circle"></i>
                        {{% else %}}
                            <i class="fas fa-info-circle"></i>
                        {{% endif %}}
                        <span>{{{{ message }}}}</span>
                        <button class="flash-close" onclick="this.parentElement.remove()">
                            <i class="fas fa-times"></i>
                        </button>
                    </div>
                {{% endfor %}}
            {{% endif %}}
        {{% endwith %}}
    </div>
    
    <div class="search-container">
        <div class="search-card">
            <div class="search-header">
                <div class="search-icon">
                    <i class="fas fa-boxes"></i>
                </div>
                <div class="search-title">Accessories Challan</div>
                <div class="search-subtitle">Enter booking reference to continue</div>
            </div>
            
            <form action="/admin/accessories/input" method="post">
                <div class="input-group">
                    <label><i class="fas fa-search" style="margin-right: 5px;"></i> BOOKING REFERENCE</label>
                    <input type="text" name="ref_no" required placeholder="e.g. IB-12345" autocomplete="off">
                </div>
                <button type="submit" style="background: linear-gradient(135deg, #8B5CF6 0%, #A78BFA 100%);">
                    Proceed to Entry <i class="fas fa-arrow-right" style="margin-left: 10px;"></i>
                </button>
            </form>
            
            <!-- History Section -->
            <div class="history-section">
                <div class="history-toggle" onclick="toggleHistory()">
                    <div class="history-toggle-left">
                        <div class="history-toggle-icon">
                            <i class="fas fa-history"></i>
                        </div>
                        <div>
                            <div class="history-toggle-text">Challan History</div>
                            <div class="history-toggle-sub">View all saved bookings</div>
                        </div>
                    </div>
                    <div class="history-badge">{{{{ history_count }}}}</div>
                </div>
                
                <div class="history-dropdown" id="historyDropdown">
                    {{% if history_bookings %}}
                        {{% for booking in history_bookings %}}
                        <a href="/admin/accessories/input_direct?ref={{{{ booking.ref }}}}" class="history-booking-item">
                            <div class="booking-item-left">
                                <div class="booking-ref">{{{{ booking.ref }}}}</div>
                                <div class="booking-info">{{{{ booking.buyer }}}} - {{{{ booking.style }}}}</div>
                            </div>
                            <div class="booking-stats">
                                <div class="booking-stat">
                                    <div class="booking-stat-value">{{{{ booking.challan_count }}}}</div>
                                    <div class="booking-stat-label">Challans</div>
                                </div>
                                <div class="booking-stat">
                                    <div class="booking-stat-value">{{{{ booking.total_qty }}}}</div>
                                    <div class="booking-stat-label">Total Qty</div>
                                </div>
                                <div class="booking-arrow">
                                    <i class="fas fa-chevron-right"></i>
                                </div>
                            </div>
                        </a>
                        {{% endfor %}}
                    {{% else %}}
                        <div class="empty-history">
                            <i class="fas fa-folder-open"></i>
                            <div>No saved bookings yet</div>
                        </div>
                    {{% endif %}}
                </div>
            </div>
            
            <div class="nav-links">
                <a href="/"><i class="fas fa-arrow-left"></i> Back to Dashboard</a>
                <a href="/logout" class="logout">Sign Out <i class="fas fa-sign-out-alt"></i></a>
            </div>
        </div>
        
        <div style="text-align: center; margin-top: 25px; color: var(--text-secondary); font-size: 11px; opacity: 0.4;">
            © 2025 Mehedi Hasan
        </div>
    </div>
    
    <script>
        function toggleHistory() {{
            document.getElementById('historyDropdown').classList.toggle('active');
        }}

        setTimeout(function() {{
            document.querySelectorAll('.flash-message').forEach(function(msg) {{
                msg.style.opacity = '0';
                setTimeout(function() {{ msg.remove(); }}, 500);
            }});
        }}, 5000);
    </script>
</body>
</html>
"""

# ==============================================================================
# ACCESSORIES INPUT TEMPLATE
# ==============================================================================

ACCESSORIES_INPUT_TEMPLATE = f"""
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Accessories Entry - MNM Software</title>
    {COMMON_STYLES}
    <style>
        .ref-badge {{
            display: inline-flex;
            align-items: center;
            gap: 10px;
            background: rgba(255, 122, 0, 0.1);
            border: 1px solid rgba(255, 122, 0, 0.2);
            padding: 10px 20px;
            border-radius: 12px;
            margin-top: 10px;
        }}
        .ref-badge .ref-no {{ font-size: 18px; font-weight: 800; color: var(--accent-orange); }}
        .ref-badge .ref-info {{ color: var(--text-secondary); font-size: 13px; font-weight: 500; }}
        .history-scroll {{ max-height: 500px; overflow-y: auto; padding-right: 5px; }}
        .challan-row {{
            display: grid;
            grid-template-columns: 60px 1fr 80px 60px 80px;
            gap: 10px;
            padding: 14px;
            background: rgba(255, 255, 255, 0.02);
            border-radius: 10px;
            margin-bottom: 8px;
            align-items: center;
            transition: var(--transition-smooth);
            border: 1px solid transparent;
        }}
        .challan-row:hover {{ background: rgba(255, 122, 0, 0.05); border-color: var(--border-glow); }}
        .line-badge {{ background: var(--gradient-orange); color: white; padding: 6px 12px; border-radius: 8px; font-weight: 700; font-size: 13px; text-align: center; }}
        .qty-value {{ font-size: 18px; font-weight: 800; color: var(--accent-green); }}
        .status-check {{ color: var(--accent-green); font-size: 20px; }}
        .print-btn {{ background: linear-gradient(135deg, #10B981 0%, #34D399 100%) !important; }}
        .refresh-btn {{ background: linear-gradient(135deg, #3B82F6 0%, #60A5FA 100%) !important; }}
        .delete-booking-btn {{ background: linear-gradient(135deg, #EF4444 0%, #F87171 100%) !important; }}
        .empty-state {{ text-align: center; padding: 50px 20px; color: var(--text-secondary); }}
        .empty-state i {{ font-size: 50px; opacity: 0.2; margin-bottom: 15px; }}
        .grid-2-cols {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
        .count-badge {{ background: var(--accent-purple); color: white; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 700; margin-left: 10px; }}
        select {{ background-color: #1a1a25 !important; color: white !important; }}
        select option {{ background-color: #1a1a25 !important; color: white !important; }}
        .header-section .btn-group {{ display: flex; align-items: center; gap: 15px; }}
    </style>
</head>
<body>
    <div class="animated-bg"></div>
    
    <div id="loading-overlay">
        <div class="spinner-container">
            <div class="spinner" id="spinner-anim"></div>
            <div class="spinner-inner"></div>
        </div>
        <div class="checkmark-container" id="success-anim">
            <div class="checkmark-circle"></div>
            <div class="anim-text">Saved!</div>
        </div>
        <div class="loading-text" id="loading-text">Saving Entry...</div>
    </div>

    <div id="flash-container">
        {{% with messages = get_flashed_messages(with_categories=true) %}}
            {{% if messages %}}
                {{% for category, message in messages %}}
                    <div class="flash-message flash-{{{{ category if category != 'message' else 'info' }}}}" role="alert">
                        {{% if category == 'success' %}}<i class="fas fa-check-circle"></i>
                        {{% elif category == 'error' %}}<i class="fas fa-times-circle"></i>
                        {{% else %}}<i class="fas fa-info-circle"></i>{{% endif %}}
                        <span>{{{{ message }}}}</span>
                        <button class="flash-close" onclick="this.parentElement.remove()"><i class="fas fa-times"></i></button>
                    </div>
                {{% endfor %}}
            {{% endif %}}
        {{% endwith %}}
    </div>

    <div class="sidebar">
        <div class="brand-logo"><i class="fas fa-boxes"></i> Accessories</div>
        <div class="nav-menu">
            <a href="/" class="nav-link"><i class="fas fa-home"></i> Dashboard</a>
            <a href="/admin/accessories" class="nav-link active"><i class="fas fa-search"></i> Search</a>
            <a href="/logout" class="nav-link" style="color: var(--accent-red); margin-top: 20px;"><i class="fas fa-sign-out-alt"></i> Sign Out</a>
        </div>
        <div class="sidebar-footer">© 2025 Mehedi Hasan</div>
    </div>
    
    <div class="main-content">
        <div class="header-section">
            <div>
                <div class="page-title">Accessories Entry</div>
                <div class="ref-badge">
                    <span class="ref-no">{{{{ ref }}}}</span>
                    <span class="ref-info">{{{{ buyer }}}} - {{{{ style }}}}</span>
                </div>
            </div>
            <div class="btn-group">
                <a href="/admin/accessories/refresh?ref={{{{ ref }}}}"><button class="refresh-btn" style="width: auto; padding: 14px 20px;"><i class="fas fa-sync-alt"></i></button></a>
                <a href="/admin/accessories/print?ref={{{{ ref }}}}" target="_blank"><button class="print-btn" style="width: auto; padding: 14px 20px;"><i class="fas fa-print"></i></button></a>
                {{% if session.role == 'admin' %}}
                <a href="/admin/accessories/delete_booking?ref={{{{ ref }}}}" onclick="return confirm('Delete Entire Booking?');"><button class="delete-booking-btn" style="width: auto; padding: 14px 20px;"><i class="fas fa-trash-alt"></i></button></a>
                {{% endif %}}
            </div>
        </div>

        <div class="dashboard-grid-2">
            <div class="card">
                <div class="section-header"><span><i class="fas fa-plus-circle" style="margin-right: 10px; color: var(--accent-orange);"></i>New Challan Entry</span></div>
                <form action="/admin/accessories/save" method="post" onsubmit="return showLoading()">
                    <input type="hidden" name="ref" value="{{{{ ref }}}}">
                    <div class="grid-2-cols">
                        <div class="input-group">
                            <label><i class="fas fa-tag" style="margin-right: 5px;"></i> TYPE</label>
                            <select name="item_type"><option value="Top">Top</option><option value="Bottom">Bottom</option></select>
                        </div>
                        <div class="input-group">
                            <label><i class="fas fa-palette" style="margin-right: 5px;"></i> COLOR</label>
                            <select name="color" required>
                                <option value="" disabled selected>Select Color</option>
                                {{% for c in colors %}}<option value="{{{{ c }}}}">{{{{ c }}}}</option>{{% endfor %}}
                            </select>
                        </div>
                    </div>
                    <div class="grid-2-cols">
                        <div class="input-group">
                            <label><i class="fas fa-industry" style="margin-right: 5px;"></i> LINE NO</label>
                            <input type="text" name="line_no" required placeholder="e.g. L-01">
                        </div>
                        <div class="input-group">
                            <label><i class="fas fa-ruler" style="margin-right: 5px;"></i> SIZE</label>
                            <input type="text" name="size" value="ALL" placeholder="Size">
                        </div>
                    </div>
                    <div class="input-group">
                        <label><i class="fas fa-sort-numeric-up" style="margin-right: 5px;"></i> QUANTITY</label>
                        <input type="number" name="qty" required placeholder="Enter Quantity" min="1">
                    </div>
                    <button type="submit"><i class="fas fa-save" style="margin-right: 10px;"></i> Save Entry</button>
                </form>
            </div>

            <div class="card">
                <div class="section-header"><span>Recent History</span><span class="count-badge">{{{{ challans|length }}}}</span></div>
                <div class="history-scroll">
                    {{% if challans %}}
                        {{% for item in challans|reverse %}}
                        <div class="challan-row" style="animation: fadeInUp 0.3s ease-out {{{{ loop.index * 0.05 }}}}s backwards;">
                            <div class="line-badge">{{{{ item.line }}}}</div>
                            <div style="color: white; font-weight: 500; font-size: 13px;">{{{{ item.color }}}}</div>
                            <div class="qty-value">{{{{ item.qty }}}}</div>
                            <div class="status-check">{{{{ item.status if item.status else '●' }}}}</div>
                            <div class="action-cell">
                                {{% if session.role == 'admin' %}}
                                <a href="/admin/accessories/edit?ref={{{{ ref }}}}&index={{{{ (challans|length) - loop.index }}}}" class="action-btn btn-edit"><i class="fas fa-pen"></i></a>
                                <form action="/admin/accessories/delete" method="POST" style="display: inline;" onsubmit="return confirm('Delete this entry?');">
                                    <input type="hidden" name="ref" value="{{{{ ref }}}}">
                                    <input type="hidden" name="index" value="{{{{ (challans|length) - loop.index }}}}">
                                    <button type="submit" class="action-btn btn-del"><i class="fas fa-trash"></i></button>
                                </form>
                                {{% else %}}
                                <span style="font-size: 10px; color: var(--text-secondary); opacity: 0.5;">🔒</span>
                                {{% endif %}}
                            </div>
                        </div>
                        {{% endfor %}}
                    {{% else %}}
                        <div class="empty-state">
                            <i class="fas fa-inbox"></i>
                            <div>No challans added yet</div>
                            <div style="font-size: 12px; margin-top: 5px;">Add your first entry using the form</div>
                        </div>
                    {{% endif %}}
                </div>
            </div>
        </div>
    </div>
    
    <script>
        function showLoading() {{
            document.getElementById('loading-overlay').style.display = 'flex';
            document.getElementById('spinner-anim').parentElement.style.display = 'block';
            document.getElementById('success-anim').style.display = 'none';
            document.getElementById('loading-text').style.display = 'block';
            return true;
        }}
        
        setTimeout(function() {{
            document.querySelectorAll('.flash-message').forEach(function(msg) {{
                msg.style.opacity = '0';
                setTimeout(function() {{ msg.remove(); }}, 500);
            }});
        }}, 5000);
    </script>
</body>
</html>
"""

# ==============================================================================
# ACCESSORIES EDIT TEMPLATE
# ==============================================================================

ACCESSORIES_EDIT_TEMPLATE = f"""
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Edit Entry - MNM Software</title>
    {COMMON_STYLES}
    <style>
        body {{ justify-content: center; align-items: center; }}
        .edit-container {{ position: relative; z-index: 10; width: 100%; max-width: 450px; padding: 20px; }}
        .edit-card {{
            background: var(--gradient-card);
            border: 1px solid var(--border-color);
            border-radius: 24px;
            padding: 45px;
            backdrop-filter: blur(20px);
            box-shadow: 0 25px 80px rgba(0, 0, 0, 0.5);
            animation: cardAppear 0.5s ease-out;
        }}
        @keyframes cardAppear {{ from {{ opacity: 0; transform: scale(0.95); }} to {{ opacity: 1; transform: scale(1); }} }}
        .edit-header {{ text-align: center; margin-bottom: 35px; }}
        .edit-icon {{
            width: 70px; height: 70px;
            background: linear-gradient(145deg, rgba(139, 92, 246, 0.2), rgba(139, 92, 246, 0.05));
            border-radius: 16px;
            display: inline-flex; align-items: center; justify-content: center;
            font-size: 28px; color: var(--accent-purple); margin-bottom: 15px;
        }}
        .edit-title {{ font-size: 24px; font-weight: 800; color: white; }}
        .cancel-link {{
            display: block; text-align: center; margin-top: 20px;
            color: var(--text-secondary); font-size: 13px; text-decoration: none;
            transition: var(--transition-smooth);
        }}
        .cancel-link:hover {{ color: var(--accent-orange); }}
    </style>
</head>
<body>
    <div class="animated-bg"></div>
    <div class="edit-container">
        <div class="edit-card">
            <div class="edit-header">
                <div class="edit-icon"><i class="fas fa-edit"></i></div>
                <div class="edit-title">Edit Entry</div>
            </div>
            <form action="/admin/accessories/update" method="post">
                <input type="hidden" name="ref" value="{{{{ ref }}}}">
                <input type="hidden" name="index" value="{{{{ index }}}}">
                <div class="input-group">
                    <label><i class="fas fa-industry" style="margin-right: 5px;"></i> LINE NO</label>
                    <input type="text" name="line_no" value="{{{{ item.line }}}}" required>
                </div>
                <div class="input-group">
                    <label><i class="fas fa-palette" style="margin-right: 5px;"></i> COLOR</label>
                    <input type="text" name="color" value="{{{{ item.color }}}}" required>
                </div>
                <div class="input-group">
                    <label><i class="fas fa-ruler" style="margin-right: 5px;"></i> SIZE</label>
                    <input type="text" name="size" value="{{{{ item.size }}}}" required>
                </div>
                <div class="input-group">
                    <label><i class="fas fa-sort-numeric-up" style="margin-right: 5px;"></i> QUANTITY</label>
                    <input type="number" name="qty" value="{{{{ item.qty }}}}" required>
                </div>
                <button type="submit" style="background: linear-gradient(135deg, #8B5CF6 0%, #A78BFA 100%);">
                    <i class="fas fa-sync-alt" style="margin-right: 10px;"></i> Update Entry
                </button>
            </form>
            <a href="/admin/accessories/input_direct?ref={{{{ ref }}}}" class="cancel-link">
                <i class="fas fa-times" style="margin-right: 5px;"></i> Cancel
            </a>
        </div>
    </div>
</body>
</html>
"""
# ==============================================================================
# CLOSING REPORT PREVIEW TEMPLATE (মূল কোডের মতোই - কোন পরিবর্তন নেই)
# ==============================================================================

CLOSING_REPORT_PREVIEW_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Closing Report Preview</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        body { background-color: #f8f9fa; padding: 30px 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; font-size: 1.1rem; }
        .container { max-width: 1400px; }
        .company-header { text-align: center; margin-bottom: 20px; border-bottom: 2px solid #000; padding-bottom: 10px; }
        .company-name { font-size: 2.2rem; font-weight: 800; color: #2c3e50; text-transform: uppercase; letter-spacing: 1px; line-height: 1; }
        .report-title { font-size: 1.1rem; color: #555; font-weight: 600; text-transform: uppercase; margin-top: 5px; }
        .date-section { font-size: 1.2rem; font-weight: 800; color: #000; margin-top: 5px; }
        .info-container { margin-bottom: 15px; background: white; padding: 15px; display: flex; justify-content: space-between; align-items: flex-end;}
        .info-row { display: flex; flex-direction: column; gap: 5px; }
        .info-item { font-size: 1.2rem; font-weight: 600; color: #444; }
        .info-label { font-weight: 800; color: #444; width: 90px; display: inline-block; }
        .info-value { font-weight: 800; color: #000; }
        .booking-box { background: #2c3e50; color: white; padding: 10px 20px; border-radius: 5px; text-align: right; box-shadow: 0 4px 10px rgba(44, 62, 80, 0.3); display: flex; flex-direction: column; align-items: flex-end; }
        .booking-label { font-size: 1.1rem; opacity: 0.9; text-transform: uppercase; letter-spacing: 1px; font-weight: 700; }
        .booking-value { font-size: 1.8rem; font-weight: 800; line-height: 1.1; }
        .table-card { background: white; border-radius: 0; margin-bottom: 30px; border: none; }
        .color-header { background-color: #2c3e50 !important; color: white; padding: 10px 15px; font-size: 1.4rem; font-weight: 800; text-transform: uppercase; border: 1px solid #000;}
        .table { margin-bottom: 0; width: 100%; border-collapse: collapse; font-size: 1rem; }
        .table th { background-color: #fff !important; color: #000 !important; text-align: center; border: 1px solid #000; padding: 8px; vertical-align: middle; font-weight: 900; font-size: 1.2rem; }
        .table td { text-align: center; vertical-align: middle; border: 1px solid #000; padding: 6px; color: #000; font-weight: 600; font-size: 1.1rem; }
        .col-3pct { background-color: #B9C2DF !important; font-weight: 700; }
        .col-input { background-color: #C4D09D !important; font-weight: 700; }
        .col-balance { font-weight: 700; color: #c0392b; }
        .total-row td { background-color: #fff !important; color: #000 !important; font-weight: 900; font-size: 1.2rem; border-top: 2px solid #000; }
        .action-bar { margin-bottom: 20px; display: flex; justify-content: flex-end; gap: 15px; position: sticky; top: 0; z-index: 1000; background: #f8f9fa; padding: 10px 0; }
        .btn-print { background-color: #2c3e50; color: white; border-radius: 50px; padding: 10px 30px; font-weight: 600; border: none; }
        .btn-excel { background-color: #27ae60; color: white; border-radius: 50px; padding: 10px 30px; font-weight: 600; text-decoration: none; display: inline-block; }
        .btn-excel:hover { color: white; background-color: #219150; }
        .footer-credit { text-align: center; margin-top: 40px; margin-bottom: 20px; font-size: 1rem; color: #2c3e50; padding-top: 10px; border-top: 1px solid #000; font-weight: 600;}
        @media print {
            @page { margin: 5mm; size: portrait; }
            body { background-color: white; padding: 0; }
            .container { max-width: 100% !important; width: 100%; }
            .no-print { display: none !important; }
            .action-bar { display: none; }
            .table th, .table td { border: 1px solid #000 !important; }
            .color-header { background-color: #2c3e50 !important; -webkit-print-color-adjust: exact; color: white !important;}
            .col-3pct { background-color: #B9C2DF !important; -webkit-print-color-adjust: exact; }
            .col-input { background-color: #C4D09D !important; -webkit-print-color-adjust: exact; }
            .booking-box { background-color: #2c3e50 !important; -webkit-print-color-adjust: exact; color: white !important; border: 1px solid #000;}
            .total-row td { font-weight: 900 !important; color: #000 !important; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="action-bar no-print">
            <a href="/" class="btn btn-outline-secondary rounded-pill px-4">Back to Dashboard</a>
            <button onclick="window.print()" class="btn btn-print"><i class="fas fa-print"></i> Print Report</button>
            <a href="/download-closing-excel?ref_no={{ ref_no }}" class="btn btn-excel"><i class="fas fa-file-excel"></i> Download Excel</a>
        </div>
        <div class="company-header">
            <div class="company-name">COTTON CLOTHING BD LTD</div>
            <div class="report-title">CLOSING REPORT [ INPUT SECTION ]</div>
            <div class="date-section">Date: <span id="date"></span></div>
        </div>
        {% if report_data %}
        <div class="info-container">
            <div class="info-row">
                <div class="info-item"><span class="info-label">Buyer:</span> <span class="info-value">{{ report_data[0].buyer }}</span></div>
                <div class="info-item"><span class="info-label">Style:</span> <span class="info-value">{{ report_data[0].style }}</span></div>
            </div>
            <div class="booking-box">
                <div class="booking-label">IR/IB NO</div>
                <div class="booking-value">{{ ref_no }}</div>
            </div>
        </div>

        {% for block in report_data %}
        <div class="table-card">
            <div class="color-header">{{ block.color }}</div>
            <table class="table">
                <thead>
                    <tr>
                        <th>Size</th>
                        <th>Actual Qty</th>
                        <th class="col-3pct">Order 3%</th>
                        <th>Cutting QC</th>
                        <th class="col-input">Input Qty</th>
                        <th>Balance</th>
                        <th>Short/Plus</th>
                        <th>Percentage</th>
                    </tr>
                </thead>
                <tbody>
                    {% for i in range(block.headers|length) %}
                    <tr>
                        <td>{{ block.headers[i] }}</td>
                        <td>{{ block.gmts_qty[i] }}</td>
                        <td class="col-3pct">{{ block.plus_3_percent[i] }}</td>
                        <td>{{ block.cutting_qc[i] if block.cutting_qc and i < block.cutting_qc|length else '0' }}</td>
                        <td class="col-input">{{ block.sewing_input[i] if block.sewing_input and i < block.sewing_input|length else '0' }}</td>
                        {% set cutting_val = (block.cutting_qc[i]|replace(',','')|int if block.cutting_qc and i < block.cutting_qc|length else 0) %}
                        {% set input_val = (block.sewing_input[i]|replace(',','')|int if block.sewing_input and i < block.sewing_input|length else 0) %}
                        {% set order_val = (block.plus_3_percent[i]|replace(',','')|int if block.plus_3_percent else 0) %}
                        <td class="col-balance">{{ cutting_val - input_val }}</td>
                        <td>{{ input_val - order_val }}</td>
                        {% if order_val > 0 %}
                        <td>{{ "%.2f"|format(((input_val - order_val) / order_val) * 100) }}%</td>
                        {% else %}
                        <td>0%</td>
                        {% endif %}
                    </tr>
                    {% endfor %}
                    <tr class="total-row">
                        <td><strong>TOTAL</strong></td>
                        {% set actual_total = namespace(value=0) %}
                        {% set order_total = namespace(value=0) %}
                        {% set cutting_total = namespace(value=0) %}
                        {% set input_total = namespace(value=0) %}
                        {% for q in block.gmts_qty %}
                            {% set actual_total.value = actual_total.value + (q|replace(',','')|int) %}
                        {% endfor %}
                        {% for q in block.plus_3_percent %}
                            {% set order_total.value = order_total.value + (q|replace(',','')|int) %}
                        {% endfor %}
                        {% if block.cutting_qc %}
                            {% for q in block.cutting_qc %}
                                {% set cutting_total.value = cutting_total.value + (q|replace(',','')|int) %}
                            {% endfor %}
                        {% endif %}
                        {% if block.sewing_input %}
                            {% for q in block.sewing_input %}
                                {% set input_total.value = input_total.value + (q|replace(',','')|int) %}
                            {% endfor %}
                        {% endif %}
                        <td><strong>{{ "{:,}".format(actual_total.value) }}</strong></td>
                        <td class="col-3pct"><strong>{{ "{:,}".format(order_total.value) }}</strong></td>
                        <td><strong>{{ "{:,}".format(cutting_total.value) }}</strong></td>
                        <td class="col-input"><strong>{{ "{:,}".format(input_total.value) }}</strong></td>
                        <td class="col-balance"><strong>{{ "{:,}".format(cutting_total.value - input_total.value) }}</strong></td>
                        <td><strong>{{ "{:,}".format(input_total.value - order_total.value) }}</strong></td>
                        {% if order_total.value > 0 %}
                        <td><strong>{{ "%.2f"|format(((input_total.value - order_total.value) / order_total.value) * 100) }}%</strong></td>
                        {% else %}
                        <td><strong>0%</strong></td>
                        {% endif %}
                    </tr>
                </tbody>
            </table>
        </div>
        {% endfor %}
        {% else %}
        <div class="alert alert-danger text-center" style="font-size: 1.2rem;">
            <i class="fas fa-exclamation-triangle"></i> No data found for this reference.
        </div>
        {% endif %}
        <div class="footer-credit">
            Report Generated by <strong>Mehedi Hasan</strong> | Cotton Clothing BD Ltd
        </div>
    </div>
    <script>
        const d = new Date();
        document.getElementById('date').innerText = 
            String(d.getDate()).padStart(2,'0') + '/' + 
            String(d.getMonth()+1).padStart(2,'0') + '/' + 
            d.getFullYear();
    </script>
</body>
</html>
"""

# ==============================================================================
# ACCESSORIES PRINT TEMPLATE (মূল কোডের মতোই - কোন পরিবর্তন নেই)
# ==============================================================================

ACCESSORIES_PRINT_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Challan Report - {{ ref }}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { 
            background: #f8f9fa; 
            padding: 20px; 
            font-family: 'Segoe UI', Tahoma, sans-serif; 
        }
        .container { max-width: 900px; }
        .report-card { 
            background: white; 
            border-radius: 10px; 
            box-shadow: 0 2px 10px rgba(0,0,0,0.1); 
            overflow: hidden; 
        }
        .report-header { 
            background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%); 
            color: white; 
            padding: 25px; 
            text-align: center; 
        }
        .report-header h1 { 
            margin: 0; 
            font-size: 1.8rem; 
            font-weight: 800; 
        }
        .report-header p { 
            margin: 5px 0 0 0; 
            opacity: 0.8; 
            font-size: 0.95rem; 
        }
        .info-bar { 
            background: #ecf0f1; 
            padding: 15px 25px; 
            display: flex; 
            justify-content: space-between; 
            flex-wrap: wrap;
            gap: 10px;
        }
        .info-item { font-weight: 600; color: #2c3e50; }
        .info-label { color: #7f8c8d; font-weight: 500; }
        .table-section { padding: 20px; }
        .table { margin-bottom: 0; }
        .table thead th { 
            background: #2c3e50; 
            color: white; 
            font-weight: 700; 
            text-align: center; 
            border: none;
            padding: 12px;
        }
        .table tbody td { 
            text-align: center; 
            vertical-align: middle; 
            font-weight: 500;
            padding: 10px;
            border-color: #dee2e6;
        }
        .table tbody tr:nth-child(even) { background: #f8f9fa; }
        .table tbody tr:hover { background: #e9ecef; }
        .line-badge { 
            background: #3498db; 
            color: white; 
            padding: 4px 12px; 
            border-radius: 20px; 
            font-weight: 700;
            font-size: 0.85rem;
        }
        .qty-cell { 
            font-weight: 800; 
            color: #27ae60; 
            font-size: 1.1rem;
        }
        .total-row { 
            background: #2c3e50 !important; 
            color: white !important; 
        }
        .total-row td { 
            color: white !important; 
            font-weight: 800 !important; 
            font-size: 1.1rem;
        }
        .action-bar { 
            margin-bottom: 20px; 
            display: flex; 
            gap: 10px; 
            justify-content: flex-end; 
        }
        .btn-back { background: #95a5a6; color: white; }
        .btn-print { background: #2c3e50; color: white; }
        .footer { 
            text-align: center; 
            padding: 15px; 
            color: #7f8c8d; 
            font-size: 0.85rem;
            border-top: 1px solid #eee;
        }
        
        @media print {
            @page { margin: 10mm; }
            body { background: white; padding: 0; }
            .action-bar { display: none !important; }
            .report-card { box-shadow: none; border: 1px solid #000; }
            .table thead th { 
                background: #2c3e50 !important; 
                -webkit-print-color-adjust: exact; 
                print-color-adjust: exact;
            }
            .total-row { 
                background: #2c3e50 !important; 
                -webkit-print-color-adjust: exact; 
                print-color-adjust: exact;
            }
            .total-row td { color: white !important; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="action-bar">
            <a href="/admin/accessories/input_direct?ref={{ ref }}" class="btn btn-back rounded-pill px-4">
                <i class="fas fa-arrow-left me-2"></i>Back
            </a>
            <button onclick="window.print()" class="btn btn-print rounded-pill px-4">
                <i class="fas fa-print me-2"></i>Print
            </button>
        </div>
        
        <div class="report-card">
            <div class="report-header">
                <h1>ACCESSORIES CHALLAN REPORT</h1>
                <p>Cotton Clothing BD Limited</p>
            </div>
            
            <div class="info-bar">
                <div class="info-item">
                    <span class="info-label">Booking:</span> {{ ref }}
                </div>
                <div class="info-item">
                    <span class="info-label">Buyer:</span> {{ buyer }}
                </div>
                <div class="info-item">
                    <span class="info-label">Style:</span> {{ style }}
                </div>
                <div class="info-item">
                    <span class="info-label">Date:</span> <span id="print-date"></span>
                </div>
            </div>
            
            <div class="table-section">
                <table class="table table-bordered">
                    <thead>
                        <tr>
                            <th width="8%">SL</th>
                            <th width="12%">Line</th>
                            <th>Color</th>
                            <th width="15%">Size</th>
                            <th width="15%">Quantity</th>
                            <th width="12%">Date</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% set total_qty = namespace(value=0) %}
                        {% for item in challans %}
                        {% set total_qty.value = total_qty.value + (item.qty|int) %}
                        <tr>
                            <td>{{ loop.index }}</td>
                            <td><span class="line-badge">{{ item.line }}</span></td>
                            <td>{{ item.color }}</td>
                            <td>{{ item.size }}</td>
                            <td class="qty-cell">{{ item.qty }}</td>
                            <td>{{ item.date }}</td>
                        </tr>
                        {% endfor %}
                        <tr class="total-row">
                            <td colspan="4" style="text-align: right; padding-right: 20px;">GRAND TOTAL</td>
                            <td>{{ "{:,}".format(total_qty.value) }}</td>
                            <td></td>
                        </tr>
                    </tbody>
                </table>
            </div>
            
            <div class="footer">
                Report Generated by <strong>Mehedi Hasan</strong> | Total Entries: {{ challans|length }}
            </div>
        </div>
    </div>
    
    <script src="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/js/all.min.js"></script>
    <script>
        const d = new Date();
        document.getElementById('print-date').innerText = 
            String(d.getDate()).padStart(2,'0') + '-' + 
            String(d.getMonth()+1).padStart(2,'0') + '-' + 
            d.getFullYear();
    </script>
</body>
</html>
"""
# ==============================================================================
# FLASK ROUTES: AUTHENTICATION
# ==============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        users = load_users()
        
        if username in users and users[username]['password'] == password:
            session.permanent = True
            session['user'] = username
            session['role'] = users[username]['role']
            session['permissions'] = users[username].get('permissions', [])
            session['login_time'] = get_bd_time().isoformat()
            
            users[username]['last_login'] = get_bd_time().strftime('%d-%b-%Y %I:%M %p')
            save_users(users)
            
            flash(f'Welcome back, {username}!', 'success')
            
            if users[username]['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('user_dashboard'))
        else:
            flash('Invalid username or password!', 'error')
            return redirect(url_for('login'))
    
    return render_template_string(LOGIN_TEMPLATE)

@app.route('/logout')
def logout():
    if 'user' in session and 'login_time' in session:
        users = load_users()
        username = session['user']
        if username in users:
            try:
                login_time = datetime.fromisoformat(session['login_time'])
                now = get_bd_time()
                if login_time.tzinfo is None:
                    login_time = bd_tz.localize(login_time)
                duration = now - login_time
                mins = int(duration.total_seconds() // 60)
                secs = int(duration.total_seconds() % 60)
                users[username]['last_duration'] = f"{mins}m {secs}s"
                save_users(users)
            except:
                pass
    
    session.clear()
    flash('You have been logged out successfully!', 'info')
    return redirect(url_for('login'))

@app.route('/')
def index():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    else:
        return redirect(url_for('user_dashboard'))

# ==============================================================================
# FLASK ROUTES: ADMIN DASHBOARD
# ==============================================================================

@app.route('/admin')
def admin_dashboard():
    if 'user' not in session or session.get('role') != 'admin':
        flash('Access denied! Admin privileges required.', 'error')
        return redirect(url_for('login'))
    
    stats = get_dashboard_summary_v2()
    return render_template_string(ADMIN_DASHBOARD_TEMPLATE, stats=stats)

@app.route('/user')
def user_dashboard():
    if 'user' not in session:
        flash('Please login to continue.', 'warning')
        return redirect(url_for('login'))
    
    return render_template_string(USER_DASHBOARD_TEMPLATE)

# ==============================================================================
# FLASK ROUTES: USER MANAGEMENT API
# ==============================================================================

@app.route('/admin/get-users')
def get_users():
    if 'user' not in session or session.get('role') != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify(load_users())

@app.route('/admin/save-user', methods=['POST'])
def save_user():
    if 'user' not in session or session.get('role') != 'admin':
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    permissions = data.get('permissions', [])
    action_type = data.get('action_type', 'create')
    
    users = load_users()
    
    if action_type == 'create':
        if username in users:
            return jsonify({"status": "error", "message": "Username already exists!"})
        
        users[username] = {
            "password": password,
            "role": "user",
            "permissions": permissions,
            "created_at": get_bd_time().strftime('%d-%b-%Y'),
            "last_login": "Never",
            "last_duration": "N/A"
        }
        flash(f'User "{username}" created successfully!', 'success')
    else:
        if username not in users:
            return jsonify({"status": "error", "message": "User not found!"})
        
        users[username]['password'] = password
        users[username]['permissions'] = permissions
        flash(f'User "{username}" updated successfully!', 'success')
    
    save_users(users)
    return jsonify({"status": "success"})

@app.route('/admin/delete-user', methods=['POST'])
def delete_user():
    if 'user' not in session or session.get('role') != 'admin':
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    
    data = request.get_json()
    username = data.get('username')
    
    users = load_users()
    
    if username in users:
        if users[username].get('role') == 'admin':
            return jsonify({"status": "error", "message": "Cannot delete admin user!"})
        
        del users[username]
        save_users(users)
        flash(f'User "{username}" deleted successfully!', 'success')
        return jsonify({"status": "success"})
    
    return jsonify({"status": "error", "message": "User not found!"})

# ==============================================================================
# FLASK ROUTES: CLOSING REPORT
# ==============================================================================

@app.route('/generate-report', methods=['POST'])
def generate_report():
    if 'user' not in session:
        flash('Please login to continue.', 'warning')
        return redirect(url_for('login'))
    
    if 'closing' not in session.get('permissions', []) and session.get('role') != 'admin':
        flash('You do not have permission for Closing Reports!', 'error')
        return redirect(url_for('index'))
    
    ref_no = request.form.get('ref_no', '').strip()
    
    if not ref_no:
        flash('Please enter a valid Booking Reference!', 'warning')
        return redirect(url_for('index'))
    
    report_data = fetch_closing_report_data(ref_no)
    
    if report_data:
        update_stats(ref_no, session['user'])
        return render_template_string(
            CLOSING_REPORT_PREVIEW_TEMPLATE,
            report_data=report_data,
            ref_no=ref_no.upper()
        )
    else:
        flash(f'No data found for reference: {ref_no}', 'error')
        return redirect(url_for('index'))

@app.route('/download-closing-excel')
def download_closing_excel():
    if 'user' not in session:
        flash('Please login to continue.', 'warning')
        return redirect(url_for('login'))
    
    ref_no = request.args.get('ref_no', '').strip()
    
    if not ref_no:
        flash('Missing reference number!', 'warning')
        return redirect(url_for('index'))
    
    report_data = fetch_closing_report_data(ref_no)
    
    if report_data:
        excel_file = create_formatted_excel_report(report_data, ref_no)
        if excel_file:
            filename = f"Closing_Report_{ref_no.upper()}_{get_bd_date_str()}.xlsx"
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
    
    flash('Failed to generate Excel file!', 'error')
    return redirect(url_for('index'))

# ==============================================================================
# FLASK ROUTES: PO SHEET GENERATOR
# ==============================================================================

@app.route('/generate-po-report', methods=['POST'])
def generate_po_report():
    if 'user' not in session:
        flash('Please login to continue.', 'warning')
        return redirect(url_for('login'))
    
    if 'po_sheet' not in session.get('permissions', []) and session.get('role') != 'admin':
        flash('You do not have permission for PO Sheet!', 'error')
        return redirect(url_for('index'))
    
    if os.path.exists(UPLOAD_FOLDER):
        shutil.rmtree(UPLOAD_FOLDER)
    os.makedirs(UPLOAD_FOLDER)
    
    uploaded_files = request.files.getlist('pdf_files')
    
    if not uploaded_files or all(f.filename == '' for f in uploaded_files):
        flash('Please select at least one PDF file!', 'warning')
        return redirect(url_for('index'))
    
    all_data = []
    final_meta = {
        'buyer': 'N/A', 'booking': 'N/A', 'style': 'N/A',
        'season': 'N/A', 'dept': 'N/A', 'item': 'N/A'
    }
    
    file_count = 0
    for file in uploaded_files:
        if file.filename == '':
            continue
        
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)
        file_count += 1
        
        data, meta = extract_data_dynamic(file_path)
        
        if meta['buyer'] != 'N/A':
            final_meta = meta
        
        if data:
            all_data.extend(data)
    
    if not all_data:
        flash('No PO table data found in the uploaded files!', 'error')
        return render_template_string(PO_REPORT_TEMPLATE, tables=None, message="No PO table data found.")
    
    df = pd.DataFrame(all_data)
    df['Color'] = df['Color'].str.strip()
    df = df[df['Color'] != ""]
    unique_colors = df['Color'].unique()
    
    final_tables = []
    grand_total_qty = 0
    
    for color in unique_colors:
        color_df = df[df['Color'] == color]
        pivot = color_df.pivot_table(
            index='P.O NO', 
            columns='Size', 
            values='Quantity', 
            aggfunc='sum', 
            fill_value=0
        )
        
        existing_sizes = pivot.columns.tolist()
        sorted_sizes = sort_sizes(existing_sizes)
        pivot = pivot[sorted_sizes]
        
        pivot['Total'] = pivot.sum(axis=1)
        grand_total_qty += pivot['Total'].sum()
        
        actual_qty = pivot.sum()
        actual_qty.name = 'Actual Qty'
        
        qty_plus_3 = (actual_qty * 1.03).round().astype(int)
        qty_plus_3.name = '3% Order Qty'
        
        pivot = pd.concat([pivot, actual_qty.to_frame().T, qty_plus_3.to_frame().T])
        
        pivot = pivot.reset_index()
        pivot = pivot.rename(columns={'index': 'P.O NO'})
        pivot.columns.name = None
        
        pd.set_option('colheader_justify', 'center')
        table_html = pivot.to_html(
            classes='table table-bordered table-striped', 
            index=False, 
            border=0
        )
        
        table_html = re.sub(r'<tr>\s*<td>', '<tr><td class="order-col">', table_html)
        table_html = table_html.replace('<th>Total</th>', '<th class="total-col-header">Total</th>')
        table_html = table_html.replace('<td>Total</td>', '<td class="total-col">Total</td>')
        
        table_html = table_html.replace('<td>Actual Qty</td>', '<td class="summary-label">Actual Qty</td>')
        table_html = table_html.replace('<td>3% Order Qty</td>', '<td class="summary-label">3% Order Qty</td>')
        table_html = re.sub(
            r'<tr>\s*<td class="summary-label">', 
            '<tr class="summary-row"><td class="summary-label">', 
            table_html
        )
        
        final_tables.append({'color': color, 'table': table_html})
    
    update_po_stats(session['user'], file_count, final_meta.get('booking', 'N/A'))
    
    flash(f'Successfully processed {file_count} file(s)!', 'success')
    
    return render_template_string(
        PO_REPORT_TEMPLATE,
        tables=final_tables,
        meta=final_meta,
        grand_total=f"{int(grand_total_qty):,}"
    )

# ==============================================================================
# FLASK ROUTES: ACCESSORIES CHALLAN SYSTEM
# ==============================================================================

@app.route('/admin/accessories')
def accessories_search():
    if 'user' not in session:
        flash('Please login to continue.', 'warning')
        return redirect(url_for('login'))
    
    if 'accessories' not in session.get('permissions', []) and session.get('role') != 'admin':
        flash('You do not have permission for Accessories!', 'error')
        return redirect(url_for('index'))
    
    history_bookings = get_all_accessories_bookings()
    
    return render_template_string(
        ACCESSORIES_SEARCH_TEMPLATE,
        history_bookings=history_bookings,
        history_count=len(history_bookings)
    )

@app.route('/admin/accessories/input', methods=['POST'])
def accessories_input_post():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    ref_no = request.form.get('ref_no', '').strip().upper()
    
    if not ref_no:
        flash('Please enter a valid booking reference!', 'warning')
        return redirect(url_for('accessories_search'))
    
    return redirect(url_for('accessories_input_direct', ref=ref_no))

@app.route('/admin/accessories/input_direct')
def accessories_input_direct():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    ref = request.args.get('ref', '').strip().upper()
    
    if not ref:
        flash('Missing booking reference!', 'warning')
        return redirect(url_for('accessories_search'))
    
    db_acc = load_accessories_db()
    
    if ref not in db_acc:
        report_data = fetch_closing_report_data(ref)
        
        if report_data:
            colors = list(set([block.get('color', 'N/A') for block in report_data]))
            buyer = report_data[0].get('buyer', 'N/A')
            style = report_data[0].get('style', 'N/A')
            
            db_acc[ref] = {
                'buyer': buyer,
                'style': style,
                'colors': colors,
                'challans': [],
                'last_api_call': get_bd_time().strftime('%d-%b-%Y %I:%M %p')
            }
            save_accessories_db(db_acc)
        else:
            flash(f'No data found for reference: {ref}', 'error')
            return redirect(url_for('accessories_search'))
    
    booking_data = db_acc[ref]
    
    return render_template_string(
        ACCESSORIES_INPUT_TEMPLATE,
        ref=ref,
        buyer=booking_data.get('buyer', 'N/A'),
        style=booking_data.get('style', 'N/A'),
        colors=booking_data.get('colors', []),
        challans=booking_data.get('challans', [])
    )

@app.route('/admin/accessories/save', methods=['POST'])
def accessories_save():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    ref = request.form.get('ref', '').strip().upper()
    item_type = request.form.get('item_type', 'Top')
    color = request.form.get('color', '')
    line_no = request.form.get('line_no', '')
    size = request.form.get('size', 'ALL')
    qty = request.form.get('qty', '0')
    
    if not ref or not color or not line_no or not qty:
        flash('Please fill all required fields!', 'warning')
        return redirect(url_for('accessories_input_direct', ref=ref))
    
    db_acc = load_accessories_db()
    
    if ref not in db_acc:
        flash('Booking not found!', 'error')
        return redirect(url_for('accessories_search'))
    
    new_challan = {
        'type': item_type,
        'color': color,
        'line': line_no,
        'size': size,
        'qty': qty,
        'date': get_bd_date_str(),
        'time': get_bd_time().strftime('%I:%M %p'),
        'user': session['user'],
        'status': '✓'
    }
    
    db_acc[ref]['challans'].append(new_challan)
    save_accessories_db(db_acc)
    
    update_accessories_stats(ref, session['user'], f"Line: {line_no}, Qty: {qty}")
    
    flash(f'Challan entry saved! Line: {line_no}, Qty: {qty}', 'success')
    return redirect(url_for('accessories_input_direct', ref=ref))

@app.route('/admin/accessories/edit')
def accessories_edit():
    if 'user' not in session or session.get('role') != 'admin':
        flash('Admin access required!', 'error')
        return redirect(url_for('login'))
    
    ref = request.args.get('ref', '').strip().upper()
    index = request.args.get('index', type=int)
    
    if not ref or index is None:
        flash('Invalid request!', 'warning')
        return redirect(url_for('accessories_search'))
    
    db_acc = load_accessories_db()
    
    if ref not in db_acc:
        flash('Booking not found!', 'error')
        return redirect(url_for('accessories_search'))
    
    challans = db_acc[ref].get('challans', [])
    
    if index < 0 or index >= len(challans):
        flash('Entry not found!', 'error')
        return redirect(url_for('accessories_input_direct', ref=ref))
    
    item = challans[index]
    
    return render_template_string(
        ACCESSORIES_EDIT_TEMPLATE,
        ref=ref,
        index=index,
        item=item
    )

@app.route('/admin/accessories/update', methods=['POST'])
def accessories_update():
    if 'user' not in session or session.get('role') != 'admin':
        flash('Admin access required!', 'error')
        return redirect(url_for('login'))
    
    ref = request.form.get('ref', '').strip().upper()
    index = request.form.get('index', type=int)
    line_no = request.form.get('line_no', '')
    color = request.form.get('color', '')
    size = request.form.get('size', 'ALL')
    qty = request.form.get('qty', '0')
    
    if not ref or index is None:
        flash('Invalid request!', 'warning')
        return redirect(url_for('accessories_search'))
    
    db_acc = load_accessories_db()
    
    if ref not in db_acc:
        flash('Booking not found!', 'error')
        return redirect(url_for('accessories_search'))
    
    challans = db_acc[ref].get('challans', [])
    
    if index < 0 or index >= len(challans):
        flash('Entry not found!', 'error')
        return redirect(url_for('accessories_input_direct', ref=ref))
    
    db_acc[ref]['challans'][index]['line'] = line_no
    db_acc[ref]['challans'][index]['color'] = color
    db_acc[ref]['challans'][index]['size'] = size
    db_acc[ref]['challans'][index]['qty'] = qty
    db_acc[ref]['challans'][index]['updated_by'] = session['user']
    db_acc[ref]['challans'][index]['updated_at'] = get_bd_time().strftime('%d-%b-%Y %I:%M %p')
    
    save_accessories_db(db_acc)
    
    flash('Entry updated successfully!', 'success')
    return redirect(url_for('accessories_input_direct', ref=ref))

@app.route('/admin/accessories/delete', methods=['POST'])
def accessories_delete():
    if 'user' not in session or session.get('role') != 'admin':
        flash('Admin access required!', 'error')
        return redirect(url_for('login'))
    
    ref = request.form.get('ref', '').strip().upper()
    index = request.form.get('index', type=int)
    
    if not ref or index is None:
        flash('Invalid request!', 'warning')
        return redirect(url_for('accessories_search'))
    
    db_acc = load_accessories_db()
    
    if ref not in db_acc:
        flash('Booking not found!', 'error')
        return redirect(url_for('accessories_search'))
    
    challans = db_acc[ref].get('challans', [])
    
    if index < 0 or index >= len(challans):
        flash('Entry not found!', 'error')
        return redirect(url_for('accessories_input_direct', ref=ref))
    
    deleted_item = challans.pop(index)
    db_acc[ref]['challans'] = challans
    save_accessories_db(db_acc)
    
    flash(f'Entry deleted! (Line: {deleted_item.get("line")}, Qty: {deleted_item.get("qty")})', 'success')
    return redirect(url_for('accessories_input_direct', ref=ref))

@app.route('/admin/accessories/delete_booking')
def accessories_delete_booking():
    if 'user' not in session or session.get('role') != 'admin':
        flash('Admin access required!', 'error')
        return redirect(url_for('login'))
    
    ref = request.args.get('ref', '').strip().upper()
    
    if not ref:
        flash('Missing booking reference!', 'warning')
        return redirect(url_for('accessories_search'))
    
    db_acc = load_accessories_db()
    
    if ref in db_acc:
        del db_acc[ref]
        save_accessories_db(db_acc)
        flash(f'Booking "{ref}" deleted successfully!', 'success')
    else:
        flash('Booking not found!', 'error')
    
    return redirect(url_for('accessories_search'))

@app.route('/admin/accessories/refresh')
def accessories_refresh():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    ref = request.args.get('ref', '').strip().upper()
    
    if not ref:
        flash('Missing booking reference!', 'warning')
        return redirect(url_for('accessories_search'))
    
    db_acc = load_accessories_db()
    
    report_data = fetch_closing_report_data(ref)
    
    if report_data:
        colors = list(set([block.get('color', 'N/A') for block in report_data]))
        buyer = report_data[0].get('buyer', 'N/A')
        style = report_data[0].get('style', 'N/A')
        
        if ref in db_acc:
            db_acc[ref]['buyer'] = buyer
            db_acc[ref]['style'] = style
            db_acc[ref]['colors'] = colors
            db_acc[ref]['last_api_call'] = get_bd_time().strftime('%d-%b-%Y %I:%M %p')
        else:
            db_acc[ref] = {
                'buyer': buyer,
                'style': style,
                'colors': colors,
                'challans': [],
                'last_api_call': get_bd_time().strftime('%d-%b-%Y %I:%M %p')
            }
        
        save_accessories_db(db_acc)
        flash('Data refreshed from server!', 'success')
    else:
        flash('Could not refresh data from server!', 'warning')
    
    return redirect(url_for('accessories_input_direct', ref=ref))

@app.route('/admin/accessories/print')
def accessories_print():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    ref = request.args.get('ref', '').strip().upper()
    
    if not ref:
        flash('Missing booking reference!', 'warning')
        return redirect(url_for('accessories_search'))
    
    db_acc = load_accessories_db()
    
    if ref not in db_acc:
        flash('Booking not found!', 'error')
        return redirect(url_for('accessories_search'))
    
    booking_data = db_acc[ref]
    
    return render_template_string(
        ACCESSORIES_PRINT_TEMPLATE,
        ref=ref,
        buyer=booking_data.get('buyer', 'N/A'),
        style=booking_data.get('style', 'N/A'),
        challans=booking_data.get('challans', [])
    )

# ==============================================================================
# FLASK ROUTES: HISTORY SEARCH API
# ==============================================================================

@app.route('/api/search-history')
def api_search_history():
    if 'user' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    query = request.args.get('q', '').strip()
    filter_type = request.args.get('type', 'all')
    time_range = request.args.get('range', '24h')
    
    if time_range == '24h':
        history = get_24h_history()
    else:
        stats_data = load_stats()
        history = stats_data.get('downloads', [])
    
    if filter_type != 'all':
        history = [h for h in history if h.get('type') == filter_type]
    
    if query:
        history = search_history(query, history)
    
    return jsonify({
        "results": history[:100],
        "count": len(history)
    })

# ==============================================================================
# FLASK ROUTES: HEALTH CHECK & ERROR HANDLERS
# ==============================================================================

@app.route('/health')
def health_check():
    return jsonify({
        "status": "healthy",
        "timestamp": get_bd_time().isoformat(),
        "version": "2.0.0"
    })

@app.errorhandler(404)
def page_not_found(e):
    flash('Page not found!', 'warning')
    return redirect(url_for('index'))

@app.errorhandler(500)
def internal_error(e):
    flash('Internal server error! Please try again.', 'error')
    return redirect(url_for('index'))

# ==============================================================================
# APPLICATION ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    print("=" * 50)
    print("  MNM Software - Production Server")
    print("  Version: 2.0.0")
    print("  Developer: Mehedi Hasan")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5000, debug=False)
    
